import os
import math
import re
import hmac
import sqlite3
import secrets
import logging
import requests
import psycopg2
import io
from psycopg2.extras import RealDictCursor
from functools import wraps
from datetime import datetime
from contextlib import contextmanager

from dotenv import load_dotenv
load_dotenv()

from flask import (
    Flask,
    Blueprint,
    render_template_string,
    request,
    jsonify,
    redirect,
    url_for,
    session,
    flash,
    send_file,
)
from markupsafe import escape

import hashlib as _hl

def hash_password(password: str) -> str:
    salt = secrets.token_hex(16)
    digest = _hl.sha256(f"{salt}:{password}".encode("utf-8")).hexdigest()
    return f"sha256${salt}${digest}"

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

# =========================
# LOGGING
# =========================
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(name)s: %(message)s",
)
logger = logging.getLogger("admin_panel")


# =========================
# STATUS CONSTANTS
# =========================
class Status:
    FREE = "free"
    USED = "used"
    ACTIVE = "active"
    BLOCKED = "blocked"


# =========================
# APP SETUP
# =========================
app = Flask(__name__)

_secret_key = os.getenv("SECRET_KEY", "").strip()
if not _secret_key:
    logger.warning("SECRET_KEY not set — generating random key (sessions won't survive restart)")
    _secret_key = secrets.token_hex(32)
app.secret_key = _secret_key

_admin_password = os.getenv("ADMIN_PASSWORD", "").strip()
if not _admin_password:
    raise RuntimeError("ADMIN_PASSWORD environment variable is required")
ADMIN_PASSWORD = _admin_password

ADMIN_PREFIX = "/control-7f2p-admin-91"
admin = Blueprint("admin", __name__, url_prefix=ADMIN_PREFIX)

PAGE_SIZE = 50
PAGE_SIZE_CODES = 50
MAX_ATTEMPTS = 3

# =========================
# CONFIG
# =========================
PUBLIC_BASE_URL = os.getenv("PUBLIC_BASE_URL", "https://codenext.ru")

YANDEX_OAUTH_TOKEN = os.getenv("YANDEX_OAUTH_TOKEN", "").strip()
YANDEX_CAMPAIGN_ID = os.getenv("YANDEX_CAMPAIGN_ID", "").strip()
YANDEX_BUSINESS_ID = os.getenv("YANDEX_BUSINESS_ID", "").strip()
YANDEX_BASE_URL = os.getenv(
    "YANDEX_BASE_URL", "https://api.partner.market.yandex.ru"
).strip()

YANDEX_HEADERS = {
    "Api-Key": YANDEX_OAUTH_TOKEN,
    "Content-Type": "application/json",
}

SLIP_TEXT_DEFAULT = (
    "Спасибо за покупку💖"
    "<p><strong>Чтобы активировать ключ:</strong></p>"
    "<ol>"
    "<li>Запустите Steam и войдите в свой аккаунт.</li>"
    "<li>Откройте меню «Игры» вверху клиента Steam.</li>"
    "<li>Выберите пункт «Активировать в Steam...».</li>"
    "<li>Следуйте инструкциям на экране, чтобы завершить активацию игры в Steam.</li>"
    "</ol>"
    "<p><strong>Если возникнут сложности с активацией</strong> — напишите в поддержку "
    "нашего магазина через кнопку «Чат с продавцом» в вашем заказе. "
    "Мы обязательно поможем разобраться.</p>"
    "<p><strong>Подписывайтесь на наш магазин</strong> — для подписчиков действует промокод "
    "на <strong>скидку 10%</strong>.</p>"
)

ACTIVATE_TILL = "2030-12-31"

BOTS = {
    "activate": {
        "name": "🚀 Website panel",
        "type": "postgresql",
        "db_name": "activate_db",
        "db_user": "activateuser",
        "db_password": os.getenv("ACTIVATE_DB_PASSWORD", ""),
        "db_host": os.getenv("ACTIVATE_DB_HOST", "localhost"),
        "db_port": int(os.getenv("ACTIVATE_DB_PORT", "5432")),
        "public_base_url": PUBLIC_BASE_URL,
        "products": {
            "plus_1m": "ChatGPT Plus · 1 месяц",
            "go_12m": "ChatGPT GO · 12 месяцев",
            "plus_12m": "ChatGPT Plus · 12 месяцев",
            "plus_account": "ChatGPT Plus · Аккаунт",
        },
        "tables": {
            "promo_codes": {"name": "Промокоды", "addable": True, "deletable": True},
            "activation_links": {"name": "Ссылки активации", "addable": True, "deletable": True},
            "orders": {"name": "Заказы", "addable": False, "deletable": False},
            "accounts": {"name": "Аккаунты", "addable": True, "deletable": True},
        },
    },
    "market": {
        "name": "🛒 Fast Game",
        "type": "sqlite",
        "db_path": "/opt/market-bot/keys.db",
        "tables": {
            "keys": {"name": "Ключи", "addable": True, "deletable": True},
            "skus": {"name": "SKU", "addable": True, "deletable": True},
            "processed_orders": {"name": "Обработанные заказы", "addable": False, "deletable": False},
            "hidden_skus": {"name": "Скрытые SKU", "addable": True, "deletable": True},
            "sku_groups": {"name": "Группы SKU", "addable": True, "deletable": True},
        },
    },
    "market2": {
        "name": "🛒 Code Next",
        "type": "sqlite",
        "db_path": "/opt/market-bot-2/keys.db",
        "tables": {
            "keys": {"name": "Ключи", "addable": True, "deletable": True},
            "skus": {"name": "SKU", "addable": True, "deletable": True},
            "processed_orders": {"name": "Обработанные заказы", "addable": False, "deletable": False},
            "hidden_skus": {"name": "Скрытые SKU", "addable": True, "deletable": True},
            "sku_groups": {"name": "Группы SKU", "addable": True, "deletable": True},
        },
    },
    "market3": {
        "name": "🛒 Digital Core",
        "type": "sqlite",
        "db_path": "/opt/market-bot-3/keys.db",
        "tables": {
            "keys": {"name": "Ключи", "addable": True, "deletable": True},
            "skus": {"name": "SKU", "addable": True, "deletable": True},
            "processed_orders": {"name": "Обработанные заказы", "addable": False, "deletable": False},
            "hidden_skus": {"name": "Скрытые SKU", "addable": True, "deletable": True},
            "sku_groups": {"name": "Группы SKU", "addable": True, "deletable": True},
        },
    },
    "notify": {
        "name": "📢 Notify & Follow-up",
        "type": "sqlite",
        "db_path": "/opt/market-bot/quiet_skus.db",  # основная БД
        "followup_db_path": "/opt/market-bot/followup.db",  # вторая БД
        "tables": {
            "quiet_skus": {"name": "Тихие SKU", "addable": True, "deletable": True},
        },
        "followup_tables": {
            "followup_chats": {"name": "Рассылка (Follow-up)", "addable": False, "deletable": True},
        },
         "shops": [
            {
                "name": "FastGame", "business_id": 216423711, "campaign_id": 148807719,
                "oauth_token": "ACMA:Cyzb6SHBJ4Vf2LKd95T0GC8ToHUtNQ0N5SnVrpLn:8a786f32"
            },
            {
                "name": "CodeNext", "business_id": 211374321, "campaign_id": 148690082,
                "oauth_token": "ACMA:VbjLCZoRgdxuUYOr1eGPjijxZpddY2hxcp4UBScC:780dcc48"
            },
            {
                "name": "DigitalCore", "business_id": 213822720, "campaign_id": 149040115,
                "oauth_token": "ACMA:lWGjlGvAv0hhGvPHxRSUsFEJowQt2BJYVSMYJvbG:a0fbb00a"
            },
        ],
    },
}

ID_COL_MAP = {
    "promo_codes": "code",
    "activation_links": "token",
    "accounts": "id",
    "keys": "id",
    "skus": "sku",
    "hidden_skus": "sku",
    "sku_groups": "sku",
    "orders": "order_id",
    "processed_orders": "order_id",
    "quiet_skus": "sku",
    "followup_chats": "order_id",
}

EXCEL_COL_TITLES = {
    "token": "Token",
    "product": "Тариф",
    "status": "Статус",
    "created_at": "Создан",
    "attempts": "Попытки",
    "cdk_code": "CDK код",
    "last_error": "Последняя ошибка",
    "code": "Код",
    "order_id": "ID заказа",
    "used_at": "Использован",
    "sku": "SKU",
    "license_key": "Ключ",
    "data": "Данные",
    "id": "ID",
    "title": "Название",
    "group_id": "Группа",
    "slip_text": "Текст отправки",
}


# =========================
# VALIDATION HELPERS
# =========================
def validate_bot_id(bot_id: str) -> str:
    if bot_id not in BOTS:
        raise ValueError(f"Invalid bot_id: {bot_id}")
    return bot_id


def validate_table_name(bot_id: str, table_name: str) -> str:
    if table_name not in BOTS[bot_id]["tables"]:
        raise ValueError(f"Invalid table: {table_name}")
    return table_name


def validate_column(col: str, allowed: list) -> str:
    if col not in allowed:
        raise ValueError(f"Invalid column: {col}")
    return col


def safe_int(value, default: int = 1) -> int:
    try:
        return int(value)
    except (ValueError, TypeError):
        return default


# =========================
# AUTH
# =========================
def login_required(func):
    @wraps(func)
    def wrapper(*args, **kwargs):
        if not session.get("admin_logged_in"):
            return redirect(url_for("admin.login"))
        return func(*args, **kwargs)
    return wrapper


# =========================
# DB
# =========================
@contextmanager
def get_db(bot_id: str):
    bot = BOTS[bot_id]
    conn = None
    try:
        if bot["type"] == "postgresql":
            conn = psycopg2.connect(
                dbname=bot["db_name"],
                user=bot["db_user"],
                password=bot.get("db_password", ""),
                host=bot.get("db_host", "localhost"),
                port=bot.get("db_port", 5432),
                cursor_factory=RealDictCursor,
            )
        else:
            conn = sqlite3.connect(bot["db_path"])
            conn.row_factory = sqlite3.Row
        yield conn
    finally:
        if conn is not None:
            conn.close()


def db_fetch_all(bot_id: str, query: str, params=None) -> list:
    with get_db(bot_id) as conn:
        cur = conn.cursor()
        cur.execute(query, params or ())
        rows = cur.fetchall()
        return [dict(r) for r in rows]


def db_fetch_one(bot_id: str, query: str, params=None):
    with get_db(bot_id) as conn:
        cur = conn.cursor()
        cur.execute(query, params or ())
        row = cur.fetchone()
        return dict(row) if row else None


def db_fetch_val(bot_id: str, query: str, params=None, default=0):
    with get_db(bot_id) as conn:
        cur = conn.cursor()
        cur.execute(query, params or ())
        row = cur.fetchone()
        if not row:
            return default
        if isinstance(row, dict):
            return list(row.values())[0]
        return row[0]


def db_execute(bot_id: str, query: str, params=None) -> int:
    with get_db(bot_id) as conn:
        cur = conn.cursor()
        cur.execute(query, params or ())
        conn.commit()
        return cur.rowcount


def db_execute_many(bot_id: str, query: str, params_list) -> int:
    with get_db(bot_id) as conn:
        cur = conn.cursor()
        cur.executemany(query, params_list)
        conn.commit()
        return cur.rowcount


# =========================
# FOLLOWUP DB (для notify бота)
# =========================
@contextmanager
def get_followup_db(bot_id: str):
    """Подключение к followup.db для notify бота"""
    bot = BOTS[bot_id]
    db_path = bot.get("followup_db_path")
    if not db_path:
        raise ValueError(f"No followup_db_path for bot {bot_id}")
    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row
    try:
        yield conn
    finally:
        conn.close()


def list_quiet_skus_admin(bot_id: str) -> list:
    """Список тихих SKU (для админки)"""
    return db_fetch_all(bot_id, "SELECT sku FROM quiet_skus ORDER BY sku")


def add_quiet_sku_admin(bot_id: str, sku: str):
    """Добавить тихий SKU"""
    db_execute(bot_id, "INSERT OR IGNORE INTO quiet_skus (sku) VALUES (?)", (sku,))


def remove_quiet_sku_admin(bot_id: str, sku: str) -> bool:
    """Удалить тихий SKU"""
    return db_execute(bot_id, "DELETE FROM quiet_skus WHERE sku = ?", (sku,)) > 0


def list_followup_orders_admin(bot_id: str, active_only: bool = True) -> list:
    """Список заказов в рассылке (как list_followup_orders в боте)"""
    with get_followup_db(bot_id) as conn:
        cur = conn.cursor()
        if active_only:
            cur.execute(
                """
                SELECT order_id, business_id, campaign_id, chat_id, variant, active,
                       last_message_id, last_checked_at, last_sent_date
                FROM followup_chats
                WHERE active = 1
                ORDER BY order_id DESC
                """
            )
        else:
            cur.execute(
                """
                SELECT order_id, business_id, campaign_id, chat_id, variant, active,
                       last_message_id, last_checked_at, last_sent_date
                FROM followup_chats
                ORDER BY order_id DESC
                """
            )
        return [dict(r) for r in cur.fetchall()]


def deactivate_followup_order_admin(bot_id: str, order_id: int) -> bool:
    """Убрать заказ из рассылки (как deactivate_followup_order в боте)"""
    with get_followup_db(bot_id) as conn:
        cur = conn.cursor()
        cur.execute(
            "UPDATE followup_chats SET active = 0 WHERE order_id = ?",
            (order_id,),
        )
        conn.commit()
        return cur.rowcount > 0
    
# =========================
# NOTIFY: YANDEX API HELPERS
# =========================
def notify_find_shop_for_order(order_id: int):
    """Ищет заказ по всем магазинам Notify-бота"""
    shops = BOTS["notify"]["shops"]
    for shop in shops:
        headers = {"Api-Key": shop["oauth_token"], "Content-Type": "application/json"}
        url = f"{YANDEX_BASE_URL}/v1/businesses/{shop['business_id']}/orders"
        try:
            resp = requests.post(url, headers=headers, params={"limit": 1}, json={"campaignIds": [shop["campaign_id"]], "orderIds": [order_id]}, timeout=15)
            if resp.status_code == 200:
                data = resp.json()
                if data.get("orders"):
                    return shop, data["orders"][0]
        except Exception as e:
            logger.error(f"Error finding order {order_id} in {shop['name']}: {e}")
    return None, None

def notify_create_chat_and_get_last_msg(shop: dict, order_id: int):
    """Открывает чат и получает последний ID сообщения"""
    headers = {"Api-Key": shop["oauth_token"], "Content-Type": "application/json"}
    url_new = f"{YANDEX_BASE_URL}/v2/businesses/{shop['business_id']}/chats/new"
    
    try:
        resp = requests.post(url_new, headers=headers, json={"context": {"id": order_id, "type": "ORDER"}}, timeout=10)
        if resp.status_code != 200 or resp.json().get("status") != "OK":
            return None, 0
        chat_id = int(resp.json()["result"]["chatId"])
    except Exception:
        return None, 0

    url_hist = f"{YANDEX_BASE_URL}/v2/businesses/{shop['business_id']}/chats/history"
    last_message_id = 0
    params_hist = {"chatId": chat_id, "limit": 100}
    
    try:
        while True:
            resp = requests.post(url_hist, headers=headers, params=params_hist, json={"messageIdFrom": 1}, timeout=10)
            if resp.status_code != 200 or resp.json().get("status") != "OK":
                break
            result = resp.json().get("result", {})
            for m in result.get("messages", []):
                mid = m.get("messageId")
                if isinstance(mid, int) and mid > last_message_id:
                    last_message_id = mid
            next_token = result.get("paging", {}).get("nextPageToken")
            if not next_token:
                break
            params_hist["page_token"] = next_token
    except Exception:
        pass

    return chat_id, last_message_id

def notify_send_message_to_market(shop: dict, chat_id: int, text: str):
    """Отправляет сообщение в чат"""
    headers = {"Api-Key": shop["oauth_token"], "Content-Type": "application/json"}
    url = f"{YANDEX_BASE_URL}/v2/businesses/{shop['business_id']}/chats/message"
    try:
        resp = requests.post(url, headers=headers, params={"chatId": chat_id}, json={"message": text}, timeout=10)
        resp.raise_for_status()
        return True, resp.json()
    except Exception as e:
        return False, str(e)


def get_placeholder(bot_id: str) -> str:
    return "%s" if BOTS[bot_id]["type"] == "postgresql" else "?"


def safe_get_bot_and_table(bot_id: str, table_name: str):
    if bot_id not in BOTS:
        return None, None
    bot = BOTS[bot_id]
    if table_name not in bot["tables"]:
        return bot, None
    return bot, bot["tables"][table_name]


def get_table_columns(bot_id: str, table_name: str) -> list:
    validate_table_name(bot_id, table_name)
    bot = BOTS[bot_id]
    with get_db(bot_id) as conn:
        cur = conn.cursor()
        if bot["type"] == "postgresql":
            cur.execute(f"SELECT * FROM {table_name} LIMIT 0")
            return [d.name if hasattr(d, "name") else d[0] for d in cur.description]
        else:
            cur.execute(f"PRAGMA table_info({table_name})")
            rows = cur.fetchall()
            return [r["name"] for r in rows]


def _insert_or_ignore(bot_id: str, query_sqlite: str, query_pg: str, params):
    if BOTS[bot_id]["type"] == "sqlite":
        db_execute(bot_id, query_sqlite, params)
    else:
        db_execute(bot_id, query_pg, params)


# =========================
# HELPERS: ACTIVATE
# =========================
def parse_activate_token(text: str) -> str:
    t = (text or "").strip()
    m = re.search(r"/[la]/([A-Za-z0-9_\-]+)", t)
    if m:
        return m.group(1)
    return t.split("/")[-1]


def create_activation_links(bot_id: str, product: str, count: int) -> list:
    ph = get_placeholder(bot_id)
    base_url = BOTS[bot_id].get("public_base_url", PUBLIC_BASE_URL)
    links = []

    with get_db(bot_id) as conn:
        cur = conn.cursor()
        for _ in range(count):
            token = secrets.token_urlsafe(16)
            cur.execute(
                f"INSERT INTO activation_links (token, product, status, attempts) "
                f"VALUES ({ph}, {ph}, {ph}, 0)",
                (token, product, Status.ACTIVE),
            )
            prefix = "/a/" if product == "plus_account" else "/l/"
            links.append(f"{base_url}{prefix}{token}")
        conn.commit()

    return links


def get_activate_stock(bot_id: str) -> list:
    products = BOTS[bot_id]["products"]
    result = []
    ph = get_placeholder(bot_id)

    for pid, name in products.items():
        if pid == "plus_account":
            cnt = db_fetch_val(
                bot_id,
                f"SELECT COUNT(*) FROM accounts WHERE status={ph}",
                (Status.FREE,),
            )
        else:
            cnt = db_fetch_val(
                bot_id,
                f"SELECT COUNT(*) FROM promo_codes WHERE product={ph} AND status={ph}",
                (pid, Status.FREE),
            )
        result.append({"product": pid, "name": name, "free": cnt})
    return result


def get_activate_codes_page(bot_id: str, product: str, page: int):
    ph = get_placeholder(bot_id)

    if product == "plus_account":
        total = db_fetch_val(bot_id, f"SELECT COUNT(*) FROM accounts WHERE status={ph}", (Status.FREE,))
        total_pages = max(1, math.ceil(total / PAGE_SIZE_CODES)) if total else 1
        page = max(1, min(page, total_pages))
        offset = (page - 1) * PAGE_SIZE_CODES

        if BOTS[bot_id]["type"] == "postgresql":
            rows = db_fetch_all(
                bot_id,
                f"SELECT data AS code FROM accounts WHERE status={ph} ORDER BY id LIMIT {ph} OFFSET {ph}",
                (Status.FREE, PAGE_SIZE_CODES, offset),
            )
        else:
            rows = db_fetch_all(
                bot_id,
                f"SELECT data AS code FROM accounts WHERE status={ph} ORDER BY id LIMIT {PAGE_SIZE_CODES} OFFSET {offset}",
                (Status.FREE,),
            )
        return rows, total, total_pages, page

    total = db_fetch_val(
        bot_id,
        f"SELECT COUNT(*) FROM promo_codes WHERE product={ph} AND status={ph}",
        (product, Status.FREE),
    )
    total_pages = max(1, math.ceil(total / PAGE_SIZE_CODES)) if total else 1
    page = max(1, min(page, total_pages))
    offset = (page - 1) * PAGE_SIZE_CODES

    if BOTS[bot_id]["type"] == "postgresql":
        rows = db_fetch_all(
            bot_id,
            f"SELECT code FROM promo_codes WHERE product={ph} AND status={ph} ORDER BY code LIMIT {ph} OFFSET {ph}",
            (product, Status.FREE, PAGE_SIZE_CODES, offset),
        )
    else:
        rows = db_fetch_all(
            bot_id,
            f"SELECT code FROM promo_codes WHERE product={ph} AND status={ph} ORDER BY code LIMIT {PAGE_SIZE_CODES} OFFSET {offset}",
            (product, Status.FREE),
        )
    return rows, total, total_pages, page


# =========================
# HELPERS: MARKET
# =========================
def get_market_stock(bot_id: str) -> list:
    return db_fetch_all(
        bot_id,
        """
        SELECT
            k.sku,
            COALESCE(s.title, '') AS title,
            SUM(CASE WHEN k.status='free' THEN 1 ELSE 0 END) AS free_cnt,
            SUM(CASE WHEN k.status='used' THEN 1 ELSE 0 END) AS used_cnt,
            COUNT(*) AS total_cnt
        FROM keys k
        LEFT JOIN skus s ON s.sku = k.sku
        LEFT JOIN hidden_skus h ON h.sku = k.sku
        WHERE h.sku IS NULL
        GROUP BY k.sku, s.title
        ORDER BY k.sku
        """,
    )


def list_visible_skus(bot_id: str) -> list:
    return db_fetch_all(
        bot_id,
        """
        SELECT s.sku, s.title
        FROM skus s
        LEFT JOIN hidden_skus h ON h.sku = s.sku
        WHERE h.sku IS NULL
        ORDER BY s.sku
        """,
    )


def get_free_keys_page_for_sku(bot_id: str, sku: str, page: int, page_size: int = 30):
    ph = get_placeholder(bot_id)
    total = db_fetch_val(
        bot_id,
        f"SELECT COUNT(*) FROM keys WHERE sku={ph} AND status={ph}",
        (sku, Status.FREE),
    )
    total_pages = max(1, math.ceil(total / page_size)) if total else 1
    page = max(1, min(page, total_pages))
    offset = (page - 1) * page_size

    if BOTS[bot_id]["type"] == "postgresql":
        rows = db_fetch_all(
            bot_id,
            f"SELECT id, license_key, status, order_id, used_at "
            f"FROM keys WHERE sku={ph} AND status={ph} "
            f"ORDER BY id LIMIT {ph} OFFSET {ph}",
            (sku, Status.FREE, page_size, offset),
        )
    else:
        rows = db_fetch_all(
            bot_id,
            f"SELECT id, license_key, status, order_id, used_at "
            f"FROM keys WHERE sku={ph} AND status={ph} "
            f"ORDER BY id LIMIT {page_size} OFFSET {offset}",
            (sku, Status.FREE),
        )
    return rows, total, total_pages, page


def get_sku_title(bot_id: str, sku: str):
    ph = get_placeholder(bot_id)
    row = db_fetch_one(bot_id, f"SELECT title FROM skus WHERE sku={ph}", (sku,))
    return row.get("title") if row else None


def set_sku_title(bot_id: str, sku: str, title):
    ph = get_placeholder(bot_id)
    _insert_or_ignore(
        bot_id,
        f"INSERT OR IGNORE INTO skus (sku) VALUES ({ph})",
        f"INSERT INTO skus (sku) VALUES ({ph}) ON CONFLICT DO NOTHING",
        (sku,),
    )
    db_execute(bot_id, f"UPDATE skus SET title={ph} WHERE sku={ph}", (title, sku))


def get_sku_slip_text(bot_id: str, sku: str):
    ph = get_placeholder(bot_id)
    row = db_fetch_one(bot_id, f"SELECT slip_text FROM skus WHERE sku={ph}", (sku,))
    return row.get("slip_text") if row else None


def set_sku_slip_text(bot_id: str, sku: str, text):
    ph = get_placeholder(bot_id)
    _insert_or_ignore(
        bot_id,
        f"INSERT OR IGNORE INTO skus (sku) VALUES ({ph})",
        f"INSERT INTO skus (sku) VALUES ({ph}) ON CONFLICT DO NOTHING",
        (sku,),
    )
    db_execute(bot_id, f"UPDATE skus SET slip_text={ph} WHERE sku={ph}", (text, sku))


def set_sku_group(bot_id: str, sku: str, group_id: int):
    ph = get_placeholder(bot_id)
    _insert_or_ignore(
        bot_id,
        f"INSERT OR REPLACE INTO sku_groups (sku, group_id) VALUES ({ph}, {ph})",
        f"INSERT INTO sku_groups (sku, group_id) VALUES ({ph}, {ph}) "
        f"ON CONFLICT (sku) DO UPDATE SET group_id=EXCLUDED.group_id",
        (sku, group_id),
    )


def remove_sku_group(bot_id: str, sku: str):
    ph = get_placeholder(bot_id)
    db_execute(bot_id, f"DELETE FROM sku_groups WHERE sku={ph}", (sku,))


def hide_sku(bot_id: str, sku: str):
    ph = get_placeholder(bot_id)
    _insert_or_ignore(
        bot_id,
        f"INSERT OR IGNORE INTO hidden_skus (sku) VALUES ({ph})",
        f"INSERT INTO hidden_skus (sku) VALUES ({ph}) ON CONFLICT DO NOTHING",
        (sku,),
    )


def unhide_sku(bot_id: str, sku: str):
    ph = get_placeholder(bot_id)
    db_execute(bot_id, f"DELETE FROM hidden_skus WHERE sku={ph}", (sku,))


# =========================
# YANDEX MARKET MANUAL ISSUE
# =========================
def market_api_ready() -> bool:
    return bool(
        YANDEX_OAUTH_TOKEN and YANDEX_CAMPAIGN_ID and YANDEX_BUSINESS_ID and YANDEX_BASE_URL
    )


def fetch_order_from_api(order_id: str):
    url = f"{YANDEX_BASE_URL}/v1/businesses/{YANDEX_BUSINESS_ID}/orders"
    params = {"limit": 1}
    body = {
        "campaignIds": [int(YANDEX_CAMPAIGN_ID)],
        "orderIds": [int(order_id)],
    }
    resp = requests.post(url, headers=YANDEX_HEADERS, params=params, json=body, timeout=20)
    resp.raise_for_status()
    data = resp.json()
    orders = data.get("orders", [])
    return orders[0] if orders else None


def deliver_digital_goods_manual(order_id, items_payload):
    url = (
        f"{YANDEX_BASE_URL}/v2/campaigns/{YANDEX_CAMPAIGN_ID}"
        f"/orders/{order_id}/deliverDigitalGoods"
    )
    resp = requests.post(url, headers=YANDEX_HEADERS, json={"items": items_payload}, timeout=20)
    return resp


def take_free_keys_for_sku(bot_id: str, sku: str, order_id, count: int) -> list:
    ph = get_placeholder(bot_id)
    keys = []
    now = datetime.utcnow().isoformat()
    with get_db(bot_id) as conn:
        cur = conn.cursor()
        if BOTS[bot_id]["type"] == "postgresql":
            cur.execute(
                f"SELECT id, license_key FROM keys WHERE sku={ph} AND status={ph} LIMIT {ph}",
                (sku, Status.FREE, int(count)),
            )
        else:
            cur.execute(
                f"SELECT id, license_key FROM keys WHERE sku={ph} AND status={ph} LIMIT {int(count)}",
                (sku, Status.FREE),
            )
        rows = cur.fetchall()
        for row in rows:
            if isinstance(row, dict) or hasattr(row, "keys"):
                rid = row["id"]
                lkey = row["license_key"]
            else:
                rid = row[0]
                lkey = row[1]
            keys.append(lkey)
            cur.execute(
                f"UPDATE keys SET status={ph}, order_id={ph}, used_at={ph} WHERE id={ph}",
                (Status.USED, order_id, now, rid),
            )
        conn.commit()
    return keys


def save_used_keys(bot_id: str, sku: str, order_id, keys: list):
    ph = get_placeholder(bot_id)
    now = datetime.utcnow().isoformat()
    params = [(sku, k, order_id, now) for k in keys]
    db_execute_many(
        bot_id,
        f"INSERT INTO keys (sku, license_key, status, order_id, used_at) "
        f"VALUES ({ph}, {ph}, '{Status.USED}', {ph}, {ph})",
        params,
    )


def mark_order_processed(bot_id: str, order_id):
    ph = get_placeholder(bot_id)
    _insert_or_ignore(
        bot_id,
        f"INSERT OR IGNORE INTO processed_orders (order_id) VALUES ({ph})",
        f"INSERT INTO processed_orders (order_id) VALUES ({ph}) ON CONFLICT DO NOTHING",
        (order_id,),
    )


def is_order_processed(bot_id: str, order_id) -> bool:
    ph = get_placeholder(bot_id)
    row = db_fetch_one(bot_id, f"SELECT 1 FROM processed_orders WHERE order_id={ph}", (order_id,))
    return bool(row)


# =========================
# EXCEL EXPORT
# =========================
def build_excel_file(columns, rows, products_dict=None, sheet_title="Export"):
    if not OPENPYXL_AVAILABLE:
        return None

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_title[:31]

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="111827", end_color="111827", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin", color="D1D5DB"),
        right=Side(style="thin", color="D1D5DB"),
        top=Side(style="thin", color="D1D5DB"),
        bottom=Side(style="thin", color="D1D5DB"),
    )

    status_fills = {
        "active": PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid"),
        "free": PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid"),
        "used": PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid"),
        "blocked": PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid"),
    }

    is_activation = "token" in columns and "product" in columns
    export_columns = list(columns)
    if is_activation:
        export_columns.append("link")

    for col_idx, col_name in enumerate(export_columns, 1):
        if col_name == "link":
            title = "Ссылка"
        else:
            title = EXCEL_COL_TITLES.get(col_name, col_name)
        cell = ws.cell(row=1, column=col_idx, value=title)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    for row_idx, row_data in enumerate(rows, 2):
        for col_idx, col_name in enumerate(export_columns, 1):
            if col_name == "link":
                token = row_data.get("token", "")
                product = row_data.get("product", "")
                prefix = "/a/" if product == "plus_account" else "/l/"
                value = f"{PUBLIC_BASE_URL}{prefix}{token}" if token else ""
            elif col_name == "product" and products_dict:
                value = products_dict.get(row_data.get(col_name), row_data.get(col_name, ""))
            else:
                value = row_data.get(col_name)
                if value is None:
                    value = ""
                else:
                    value = str(value)

            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center", wrap_text=False)

            if col_name == "status":
                status_val = (value or "").lower()
                if status_val in status_fills:
                    cell.fill = status_fills[status_val]

    for col_idx, col_name in enumerate(export_columns, 1):
        max_len = len(EXCEL_COL_TITLES.get(col_name, col_name))
        for row_idx in range(2, min(len(rows) + 2, 200)):
            cell_val = ws.cell(row=row_idx, column=col_idx).value
            if cell_val:
                max_len = max(max_len, len(str(cell_val)))
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max_len + 4, 60)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


# =========================
# ESCAPE HELPER
# =========================
def e(value) -> str:
    if value is None:
        return ""
    return str(escape(value))
# =========================
# UI TEMPLATES
# =========================
LOGIN_TEMPLATE = """
<!doctype html>
<html lang="ru">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>Вход</title>
<style>
body{font-family:Arial,sans-serif;background:#0f172a;color:#111;margin:0;padding:40px}
.card{max-width:420px;margin:80px auto;background:#fff;border-radius:16px;padding:28px;box-shadow:0 10px 40px rgba(0,0,0,.2)}
input,button{width:100%;padding:12px 14px;border-radius:10px;font-size:15px}
input{border:1px solid #d1d5db;margin:12px 0}
button{background:#4f46e5;color:#fff;border:none;font-weight:700;cursor:pointer}
.err{background:#fee2e2;color:#991b1b;padding:10px 12px;border-radius:10px;margin-bottom:12px}
</style>
</head>
<body>
<div class="card">
<h1>🔐 Вход в админку</h1>
<p>Введите пароль администратора</p>
{% with messages = get_flashed_messages() %}
  {% if messages %}
    {% for msg in messages %}
      <div class="err">{{ msg }}</div>
    {% endfor %}
  {% endif %}
{% endwith %}
<form method="post">
  <input type="password" name="password" required placeholder="Пароль">
  <button type="submit">Войти</button>
</form>
</div>
</body>
</html>
"""

BASE_TEMPLATE = """
<!doctype html>
<html lang="ru">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>Админка</title>
<style>
body{font-family:Arial,sans-serif;background:#f3f4f6;margin:0;color:#111}
.wrap{max-width:1500px;margin:0 auto;padding:20px}
.top{background:#111827;color:#fff;padding:18px 20px;border-radius:16px;display:flex;justify-content:space-between;align-items:center;gap:16px;flex-wrap:wrap}
.top a{color:#fff;text-decoration:none;background:#374151;padding:10px 14px;border-radius:10px}
.box{background:#fff;border-radius:16px;padding:18px;margin-top:18px;box-shadow:0 6px 24px rgba(0,0,0,.06)}
.tabs{display:flex;gap:8px;flex-wrap:wrap}
.tabs a{padding:10px 14px;border-radius:10px;text-decoration:none;background:#e5e7eb;color:#111}
.tabs a.active{background:#4f46e5;color:#fff}
.controls{display:flex;gap:10px;flex-wrap:wrap;align-items:center}
input,select,textarea,button{padding:10px 12px;border-radius:10px;font-size:14px;font-family:inherit}
input,select,textarea{border:1px solid #d1d5db;background:#fff}
button{border:none;background:#4f46e5;color:#fff;cursor:pointer;font-weight:700}
button.gray{background:#6b7280}
button.red{background:#dc2626}
button.green{background:#059669}
button.orange{background:#d97706}
button.excel{background:#217346}
.stats{display:grid;grid-template-columns:repeat(auto-fit,minmax(220px,1fr));gap:12px}
.stat{background:#111827;color:#fff;border-radius:14px;padding:16px}
.stat .v{font-size:28px;font-weight:800;margin-top:8px}
table{width:100%;border-collapse:collapse}
th,td{padding:10px 12px;border-bottom:1px solid #e5e7eb;text-align:left;vertical-align:top}
th{background:#111827;color:#fff}
.badge{display:inline-block;padding:4px 8px;border-radius:999px;font-size:12px;font-weight:700}
.free,.active{background:#d1fae5;color:#065f46}
.used{background:#fee2e2;color:#991b1b}
.blocked{background:#fef3c7;color:#92400e}
.other{background:#e5e7eb;color:#374151}
pre,code{white-space:pre-wrap;word-break:break-word}
.flash{padding:12px 14px;border-radius:10px;margin-bottom:12px}
.flash.success{background:#d1fae5;color:#065f46}
.flash.error{background:#fee2e2;color:#991b1b}
.flash.info{background:#dbeafe;color:#1e40af}
.grid2{display:grid;grid-template-columns:1fr 1fr;gap:16px}
@media(max-width:900px){.grid2{grid-template-columns:1fr}}
.small{font-size:12px;color:#6b7280}
.filter-bar{display:flex;gap:8px;flex-wrap:wrap;margin-bottom:12px;align-items:center}
.filter-bar a{padding:8px 14px;border-radius:10px;text-decoration:none;font-size:13px;font-weight:600;background:#e5e7eb;color:#374151;transition:all .15s}
.filter-bar a:hover{background:#d1d5db}
.filter-bar a.active{background:#4f46e5;color:#fff}
.filter-bar .label{font-size:13px;color:#6b7280;font-weight:600}
</style>
</head>
<body>
<div class="wrap">
  <div class="top">
    <div>
      <h1 style="margin:0">🎛️ Админ-панель</h1>
      <div class="small">Скрытый путь: {{ admin_prefix }}</div>
    </div>
    <a href="{{ url_for('admin.logout') }}">Выйти</a>
  </div>

  <div class="box">
    <div class="tabs">
      {% for bid, b in bots.items() %}
        <a class="{% if current_bot==bid %}active{% endif %}" href="{{ url_for('admin.index', bot=bid) }}">{{ b.name }}</a>
      {% endfor %}
    </div>
  </div>

  {% with messages = get_flashed_messages(with_categories=true) %}
    {% if messages %}
      <div class="box">
        {% for category, msg in messages %}
          <div class="flash {{ category }}">{{ msg }}</div>
        {% endfor %}
      </div>
    {% endif %}
  {% endwith %}

  {{ content|safe }}
</div>
</body>
</html>
"""


def render_page(content: str, **ctx):
    return render_template_string(
        BASE_TEMPLATE,
        content=content,
        admin_prefix=ADMIN_PREFIX,
        bots=BOTS,
        **ctx,
    )


def badge_html(value: str) -> str:
    v = (value or "").lower()
    cls = "other"
    if v in (Status.FREE, Status.ACTIVE):
        cls = v
    elif v == Status.USED:
        cls = "used"
    elif v == Status.BLOCKED:
        cls = "blocked"
    return f'<span class="badge {e(cls)}">{e(value)}</span>'


def _build_product_filter_bar(current_bot, current_table, q, current_product_filter):
    bot = BOTS.get(current_bot)
    if not bot:
        return ""
    products = bot.get("products")
    if not products:
        return ""

    all_active = "active" if not current_product_filter else ""
    bar = f'<div class="filter-bar"><span class="label">Тариф:</span>'
    bar += (
        f'<a class="{all_active}" '
        f'href="{url_for("admin.index", bot=current_bot, table=current_table, q=q, page=1)}">'
        f'Все</a>'
    )

    for pid, pname in products.items():
        is_active = "active" if current_product_filter == pid else ""
        bar += (
            f'<a class="{is_active}" '
            f'href="{url_for("admin.index", bot=current_bot, table=current_table, q=q, page=1, product=pid)}">'
            f'{e(pname)}</a>'
        )
    bar += '</div>'
    return bar


# =========================
# ROUTES: AUTH
# =========================
@admin.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        pw = request.form.get("password", "")
        if hmac.compare_digest(pw, ADMIN_PASSWORD):
            session["admin_logged_in"] = True
            logger.info("Admin login successful from %s", request.remote_addr)
            return redirect(url_for("admin.index"))
        logger.warning("Admin login failed from %s", request.remote_addr)
        flash("Неверный пароль")
    return render_template_string(LOGIN_TEMPLATE)


@admin.route("/logout")
def logout():
    logger.info("Admin logout from %s", request.remote_addr)
    session.clear()
    return redirect(url_for("admin.login"))


# =========================
# EXCEL EXPORT ROUTE
# =========================
@admin.route("/export-excel")
@login_required
def export_excel():
    if not OPENPYXL_AVAILABLE:
        flash("Библиотека openpyxl не установлена. Выполните: pip install openpyxl", "error")
        return redirect(request.referrer or url_for("admin.index"))

    bot_id = request.args.get("bot", "").strip()
    table_name = request.args.get("table", "").strip()
    product_filter = request.args.get("product", "").strip()
    q = request.args.get("q", "").strip()

    if bot_id not in BOTS:
        flash("Неверный bot", "error")
        return redirect(url_for("admin.index"))

    bot, table_meta = safe_get_bot_and_table(bot_id, table_name)
    if not bot or not table_meta:
        flash("Неверная таблица", "error")
        return redirect(url_for("admin.index", bot=bot_id))

    try:
        validate_table_name(bot_id, table_name)
        columns = get_table_columns(bot_id, table_name)
        ph = get_placeholder(bot_id)
        products_dict = bot.get("products", {})

        has_product_col = "product" in columns
        if product_filter and (not has_product_col or product_filter not in products_dict):
            product_filter = ""

        where_parts = []
        params = []

        if has_product_col and product_filter:
            where_parts.append(f"product={ph}")
            params.append(product_filter)

        if q and columns:
            search_parts = []
            if table_name == "activation_links":
                search_token = parse_activate_token(q)
                if bot["type"] == "postgresql":
                    search_parts.append(f"token ILIKE {ph}")
                    search_parts.append(f"cdk_code ILIKE {ph}")
                else:
                    search_parts.append(f"token LIKE {ph}")
                    search_parts.append(f"cdk_code LIKE {ph}")
                params.append(f"%{search_token}%")
                params.append(f"%{q}%")
            else:
                for col in columns:
                    validate_column(col, columns)
                    if bot["type"] == "postgresql":
                        search_parts.append(f"CAST({col} AS TEXT) ILIKE {ph}")
                    else:
                        search_parts.append(f"CAST({col} AS TEXT) LIKE {ph}")
                    params.append(f"%{q}%")
            where_parts.append("(" + " OR ".join(search_parts) + ")")

        where_clause = ("WHERE " + " AND ".join(where_parts)) if where_parts else ""

        rows = db_fetch_all(
            bot_id,
            f"SELECT * FROM {table_name} {where_clause} ORDER BY rowid"
            if BOTS[bot_id]["type"] == "sqlite"
            else f"SELECT * FROM {table_name} {where_clause}",
            tuple(params),
        )

        parts = [table_name]
        if product_filter:
            safe_product = re.sub(r'[^\w\-]', '_', product_filter)
            parts.append(safe_product)
        if q:
            parts.append("search")
        timestamp = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
        parts.append(timestamp)
        filename = "_".join(parts) + ".xlsx"

        sheet_title = table_meta["name"]
        if product_filter and products_dict:
            sheet_title += f" — {products_dict.get(product_filter, product_filter)}"

        excel_file = build_excel_file(
            columns=columns,
            rows=rows,
            products_dict=products_dict if products_dict else None,
            sheet_title=sheet_title,
        )

        if not excel_file:
            flash("Ошибка генерации Excel", "error")
            return redirect(url_for("admin.index", bot=bot_id, table=table_name))

        logger.info(
            "Excel export: bot=%s table=%s product=%s q=%s rows=%d",
            bot_id, table_name, product_filter, q, len(rows),
        )

        return send_file(
            excel_file,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name=filename,
        )

    except Exception as exc:
        logger.exception("Excel export error: %s", exc)
        flash(f"Ошибка экспорта: {exc}", "error")
        return redirect(url_for("admin.index", bot=bot_id, table=table_name))


# =========================
# ROUTES: HOME / INDEX
# =========================
@admin.route("/")
@login_required
def index():
    current_bot = request.args.get("bot", "").strip()
    current_table = request.args.get("table", "").strip()
    q = request.args.get("q", "").strip()
    page = max(1, safe_int(request.args.get("page", 1)))
    product_filter = request.args.get("product", "").strip()

    if not current_bot or current_bot not in BOTS:
        html = """
        <div class="box">
          <h2>Выбери бота</h2>
          <p>Сверху выбери нужный магазин или Activate Bot.</p>
        </div>
        """
        return render_page(html, current_bot=current_bot)

    if not current_table:
        return _render_dashboard(current_bot)
    

    return _render_table_view(current_bot, current_table, q, page, product_filter)


def _render_dashboard(current_bot: str):
    if current_bot == "notify":
        return redirect(url_for("admin.notify_dashboard"))
    
    if current_bot == "activate":
        stocks = get_activate_stock(current_bot)
        stats_html = "".join(
            [
                f'<div class="stat"><div>{e(s["name"])}</div>'
                f'<div class="v">{e(str(s["free"]))}</div>'
                f'<div class="small">Свободно</div></div>'
                for s in stocks
            ]
        )

        ph = get_placeholder(current_bot)
        cab_users = db_fetch_val(current_bot, "SELECT COUNT(*) FROM users")
        cab_wd_new = db_fetch_val(current_bot, f"SELECT COUNT(*) FROM withdraw_requests WHERE status={ph}", ("new",))
        cab_wd_new_sum = db_fetch_val(current_bot, f"SELECT COALESCE(SUM(amount),0) FROM withdraw_requests WHERE status={ph}", ("new",))
        cab_wd_paid = db_fetch_val(current_bot, f"SELECT COALESCE(SUM(amount),0) FROM withdraw_requests WHERE status={ph}", ("approved",))
        cab_ref_bonus = db_fetch_val(current_bot, f"SELECT COALESCE(SUM(amount),0) FROM wallet_transactions WHERE tx_type={ph}", ("referral_bonus",))

        cab_stats_html = f"""
        <div class="stat"><div>Пользователей</div><div class="v">{cab_users}</div></div>
        <div class="stat"><div>Новых заявок</div><div class="v">{cab_wd_new}</div><div class="small">на вывод</div></div>
        <div class="stat"><div>Ожидает выплаты</div><div class="v">{cab_wd_new_sum} ₽</div></div>
        <div class="stat"><div>Выплачено всего</div><div class="v">{cab_wd_paid} ₽</div></div>
        <div class="stat"><div>Реф. бонусов</div><div class="v">{cab_ref_bonus} ₽</div><div class="small">начислено</div></div>
        """

        table_tabs = "".join(
            [
                f'<a href="{url_for("admin.index", bot=current_bot, table=t)}">'
                f"{e(meta['name'])}</a>"
                for t, meta in BOTS[current_bot]["tables"].items()
            ]
        )

        content = f"""
        <div class="box">
          <h2>{e(BOTS[current_bot]["name"])}</h2>
          <div class="controls" style="margin-top:12px">
            <a href="{url_for('admin.activate_create_links_form', bot=current_bot)}"><button type="button" class="green">Создать ссылки</button></a>
            <a href="{url_for('admin.activate_upload_form', bot=current_bot)}"><button type="button">Загрузить лицензии</button></a>
            <a href="{url_for('admin.activate_check_link_form', bot=current_bot)}"><button type="button" class="gray">Проверить ссылку</button></a>
            <a href="{url_for('admin.activate_find_order_form', bot=current_bot)}"><button type="button" class="gray">Найти заказ</button></a>
            <a href="{url_for('admin.activate_keys_choose', bot=current_bot)}"><button type="button" class="orange">Лицензии по тарифу</button></a>
            <a href="{url_for('admin.activate_links_choose', bot=current_bot)}"><button type="button" class="orange">Ссылки по тарифу</button></a>
          </div>
          <div class="controls" style="margin-top:8px">
            <a href="{url_for('admin.activate_withdrawals', bot=current_bot)}"><button type="button" class="red">💰 Заявки на вывод{(' (' + str(cab_wd_new) + ')') if cab_wd_new else ''}</button></a>
            <a href="{url_for('admin.activate_cabinet_users', bot=current_bot)}"><button type="button" class="gray">👥 Пользователи</button></a>
            <a href="{url_for('admin.activate_referral_analytics', bot=current_bot)}"><button type="button" class="orange">📊 Реф. аналитика</button></a>
          </div>
        </div>
        <div class="box">
          <h3>Остатки по тарифам</h3>
          <div class="stats">{stats_html}</div>
        </div>
        <div class="box">
          <h3>Кабинет</h3>
          <div class="stats">{cab_stats_html}</div>
        </div>
        <div class="box">
          <h3>Таблицы</h3>
          <div class="tabs">{table_tabs}</div>
        </div>
        """
        return render_page(content, current_bot=current_bot)

    else:
        stocks = get_market_stock(current_bot)
        cards = []
        for s in stocks:
            label = (
                f'{e(s["title"])} ({e(s["sku"])})'
                if s.get("title")
                else e(s["sku"])
            )
            cards.append(
                f'<div class="stat"><div>{label}</div>'
                f'<div class="v">{e(str(s["free_cnt"]))}</div>'
                f'<div class="small">free / used: {e(str(s["used_cnt"]))}</div></div>'
            )

        table_tabs = "".join(
            [
                f'<a href="{url_for("admin.index", bot=current_bot, table=t)}">'
                f"{e(meta['name'])}</a>"
                for t, meta in BOTS[current_bot]["tables"].items()
            ]
        )

        content = f"""
        <div class="box">
          <h2>{e(BOTS[current_bot]["name"])}</h2>
          <div class="controls" style="margin-top:12px">
            <a href="{url_for('admin.market_add_sku_form', bot=current_bot)}"><button type="button" class="green">Добавить SKU</button></a>
            <a href="{url_for('admin.market_upload_keys_form', bot=current_bot)}"><button type="button">Загрузить ключи</button></a>
            <a href="{url_for('admin.market_find_order_keys_form', bot=current_bot)}"><button type="button" class="gray">Ключи по заказу</button></a>
            <a href="{url_for('admin.market_hide_show_form', bot=current_bot)}"><button type="button" class="gray">Скрыть / показать SKU</button></a>
            <a href="{url_for('admin.market_groups_form', bot=current_bot)}"><button type="button" class="orange">Группы SKU</button></a>
            <a href="{url_for('admin.market_manual_issue_form', bot=current_bot)}"><button type="button" class="red">Ручная выдача</button></a>
          </div>
        </div>
        <div class="box">
          <h3>Остатки по SKU</h3>
          <div class="stats">{''.join(cards) if cards else '<p>Нет данных</p>'}</div>
        </div>
        <div class="box">
          <h3>Таблицы</h3>
          <div class="tabs">{table_tabs}</div>
        </div>
        """
        return render_page(content, current_bot=current_bot)


def _render_table_view(current_bot: str, current_table: str, q: str, page: int, product_filter: str = ""):
    bot, table_meta = safe_get_bot_and_table(current_bot, current_table)
    if not bot or not table_meta:
        flash("Неверная таблица", "error")
        return redirect(url_for("admin.index", bot=current_bot))

    try:
        validate_table_name(current_bot, current_table)
        columns = get_table_columns(current_bot, current_table)
        ph = get_placeholder(current_bot)

        has_product_col = "product" in columns
        products_dict = bot.get("products", {})
        show_product_filter = has_product_col and products_dict and current_table in ("activation_links", "promo_codes")

        if product_filter and product_filter not in products_dict:
            product_filter = ""

        where_parts = []
        params = []

        if show_product_filter and product_filter:
            where_parts.append(f"product={ph}")
            params.append(product_filter)

        if q and columns:
            search_parts = []
            if current_table == "activation_links":
                search_token = parse_activate_token(q)
                if bot["type"] == "postgresql":
                    search_parts.append(f"token ILIKE {ph}")
                    search_parts.append(f"cdk_code ILIKE {ph}")
                else:
                    search_parts.append(f"token LIKE {ph}")
                    search_parts.append(f"cdk_code LIKE {ph}")
                params.append(f"%{search_token}%")
                params.append(f"%{q}%")
            else:
                for col in columns:
                    validate_column(col, columns)
                    if bot["type"] == "postgresql":
                        search_parts.append(f"CAST({col} AS TEXT) ILIKE {ph}")
                    else:
                        search_parts.append(f"CAST({col} AS TEXT) LIKE {ph}")
                    params.append(f"%{q}%")
            where_parts.append("(" + " OR ".join(search_parts) + ")")

        where_clause = ("WHERE " + " AND ".join(where_parts)) if where_parts else ""

        total_count = db_fetch_val(
            current_bot,
            f"SELECT COUNT(*) FROM {current_table} {where_clause}",
            tuple(params),
        )
        total_pages = max(1, math.ceil(total_count / PAGE_SIZE)) if total_count else 1
        page = max(1, min(page, total_pages))
        offset = (page - 1) * PAGE_SIZE

        if BOTS[current_bot]["type"] == "postgresql":
            rows = db_fetch_all(
                current_bot,
                f"SELECT * FROM {current_table} {where_clause} LIMIT {ph} OFFSET {ph}",
                tuple(params) + (PAGE_SIZE, offset),
            )
        else:
            rows = db_fetch_all(
                current_bot,
                f"SELECT * FROM {current_table} {where_clause} LIMIT {PAGE_SIZE} OFFSET {offset}",
                tuple(params),
            )

        stats = _build_table_stats(current_bot, current_table, columns, total_count, product_filter)
        id_col = ID_COL_MAP.get(current_table, columns[0] if columns else None)

        cards = "".join(
            [
                f'<div class="stat"><div>{e(str(t))}</div>'
                f'<div class="v">{e(str(v))}</div></div>'
                for t, v in stats
            ]
        )

        top_actions = ""
        if current_table in ("promo_codes", "keys"):
            top_actions += (
                f'<a href="{url_for("admin.bulk_add_form", bot=current_bot, table=current_table)}">'
                f'<button type="button" class="green">Массовая загрузка</button></a>'
            )
        if table_meta.get("addable"):
            top_actions += (
                f' <a href="{url_for("admin.add_record_form", bot=current_bot, table=current_table)}">'
                f'<button type="button">Добавить запись</button></a>'
            )

        excel_btn = ""
        if OPENPYXL_AVAILABLE:
            export_url = url_for(
                "admin.export_excel",
                bot=current_bot,
                table=current_table,
                product=product_filter,
                q=q,
            )
            excel_btn = (
                f' <a href="{export_url}">'
                f'<button type="button" class="excel">📥 Скачать Excel</button></a>'
            )

        search_form = f"""
        <form method="get" class="controls">
          <input type="hidden" name="bot" value="{e(current_bot)}">
          <input type="hidden" name="table" value="{e(current_table)}">
          {'<input type="hidden" name="product" value="' + e(product_filter) + '">' if product_filter else ''}
          <input type="text" name="q" value="{e(q)}" placeholder="Поиск">
          <button type="submit">Найти</button>
          <a href="{url_for('admin.index', bot=current_bot, table=current_table, product=product_filter)}"><button type="button" class="gray">Сбросить</button></a>
          {top_actions}
          {excel_btn}
        </form>
        """

        filter_bar_html = ""
        if show_product_filter:
            filter_bar_html = _build_product_filter_bar(current_bot, current_table, q, product_filter)

        if rows:
            head = "".join([f"<th>{e(c)}</th>" for c in columns])
            if table_meta.get("deletable") and id_col:
                head += "<th>Действия</th>"

            body_rows = []
            for r in rows:
                cells = []
                for c in columns:
                    val = r.get(c)
                    if c == "status":
                        cells.append(f"<td>{badge_html(str(val or '—'))}</td>")
                    elif c == "product" and products_dict:
                        pname = products_dict.get(val, val)
                        cells.append(f"<td>{e(str(pname))}</td>")
                    elif c in ("license_key", "code", "data", "token", "cdk_code", "last_error"):
                        sval = "" if val is None else str(val)
                        short = sval[:80] + ("..." if len(sval) > 80 else "")
                        cells.append(f"<td><code>{e(short)}</code></td>")
                    else:
                        cells.append(f"<td>{e(str(val)) if val not in [None, ''] else '—'}</td>")
                if table_meta.get("deletable") and id_col:
                    id_val = r.get(id_col)
                    del_form = (
                        f'<form method="post" '
                        f'action="{url_for("admin.delete_record_post")}?bot={e(current_bot)}&table={e(current_table)}" '
                        f"onsubmit=\"return confirm('Удалить запись?')\">"
                        f'<input type="hidden" name="id" value="{e(str(id_val))}">'
                        f'<button type="submit" class="red">Удалить</button>'
                        f"</form>"
                    )
                    cells.append(f"<td>{del_form}</td>")
                body_rows.append("<tr>" + "".join(cells) + "</tr>")

            pager = _build_pager(current_bot, current_table, q, page, total_pages, total_count, product_filter)

            content = f"""
            <div class="box"><h2>{e(table_meta["name"])}</h2>{filter_bar_html}{search_form}</div>
            <div class="box"><div class="stats">{cards}</div></div>
            <div class="box">
              <table>
                <thead><tr>{head}</tr></thead>
                <tbody>{''.join(body_rows)}</tbody>
              </table>
              {pager}
            </div>
            """
        else:
            content = f"""
            <div class="box"><h2>{e(table_meta["name"])}</h2>{filter_bar_html}{search_form}</div>
            <div class="box"><div class="stats">{cards}</div></div>
            <div class="box"><p>Нет данных</p></div>
            """

        return render_page(content, current_bot=current_bot)

    except Exception as exc:
        logger.exception("Error loading table %s/%s: %s", current_bot, current_table, exc)
        flash(f"Ошибка загрузки таблицы: {exc}", "error")
        return redirect(url_for("admin.index", bot=current_bot))


def _build_table_stats(current_bot, current_table, columns, total_count, product_filter=""):
    ph = get_placeholder(current_bot)

    product_where = ""
    product_params = ()
    if product_filter and "product" in columns:
        product_where = f" AND product={ph}"
        product_params = (product_filter,)

    if "status" in columns:
        if current_table == "activation_links":
            active = db_fetch_val(current_bot, f"SELECT COUNT(*) FROM activation_links WHERE status={ph}{product_where}", (Status.ACTIVE,) + product_params)
            used = db_fetch_val(current_bot, f"SELECT COUNT(*) FROM activation_links WHERE status={ph}{product_where}", (Status.USED,) + product_params)
            blocked = db_fetch_val(current_bot, f"SELECT COUNT(*) FROM activation_links WHERE status={ph}{product_where}", (Status.BLOCKED,) + product_params)
            return [("Активных", active), ("Использованных", used), ("Заблокированных", blocked)]
        else:
            free = db_fetch_val(current_bot, f"SELECT COUNT(*) FROM {current_table} WHERE status={ph}{product_where}", (Status.FREE,) + product_params)
            used = db_fetch_val(current_bot, f"SELECT COUNT(*) FROM {current_table} WHERE status={ph}{product_where}", (Status.USED,) + product_params)
            return [("Свободных", free), ("Использованных", used), ("Всего", free + used)]
    return [("Записей", total_count)]


def _build_pager(current_bot, current_table, q, page, total_pages, total_count, product_filter=""):
    pager = '<div class="controls" style="margin-top:12px">'
    if page > 1:
        pager += (
            f'<a href="{url_for("admin.index", bot=current_bot, table=current_table, q=q, page=page - 1, product=product_filter)}">'
            f'<button type="button" class="gray">← Назад</button></a>'
        )
    pager += f"<span>Страница {page} из {total_pages} | Всего: {total_count}</span>"
    if page < total_pages:
        pager += (
            f'<a href="{url_for("admin.index", bot=current_bot, table=current_table, q=q, page=page + 1, product=product_filter)}">'
            f'<button type="button" class="gray">Вперёд →</button></a>'
        )
    pager += "</div>"
    return pager


# =========================
# GENERIC FORMS
# =========================
@admin.route("/form/add")
@login_required
def add_record_form():
    bot_id = request.args.get("bot", "").strip()
    table = request.args.get("table", "").strip()
    bot, table_meta = safe_get_bot_and_table(bot_id, table)
    if not bot or not table_meta:
        flash("Неверный bot или table", "error")
        return redirect(url_for("admin.index", bot=bot_id))

    html = f'<div class="box"><h2>Добавить запись: {e(table_meta["name"])}</h2>'

    if table == "promo_codes":
        options = "".join([f'<option value="{e(k)}">{e(v)}</option>' for k, v in bot.get("products", {}).items() if k != "plus_account"])
        html += f"""
        <form method="post" action="{url_for('admin.add_record')}">
          <input type="hidden" name="bot" value="{e(bot_id)}">
          <input type="hidden" name="table" value="{e(table)}">
          <div class="controls">
            <input type="text" name="code" required placeholder="Код">
            <select name="product" required>{options}</select>
            <button type="submit">Добавить</button>
          </div>
        </form>
        """
    elif table == "accounts":
        html += f"""
        <form method="post" action="{url_for('admin.add_record')}">
          <input type="hidden" name="bot" value="{e(bot_id)}">
          <input type="hidden" name="table" value="{e(table)}">
          <textarea name="data" required placeholder="email:password"></textarea>
          <div style="margin-top:12px"><button type="submit">Добавить</button></div>
        </form>
        """
    elif table == "activation_links":
        options = "".join([f'<option value="{e(k)}">{e(v)}</option>' for k, v in bot.get("products", {}).items()])
        html += f"""
        <form method="post" action="{url_for('admin.add_record')}">
          <input type="hidden" name="bot" value="{e(bot_id)}">
          <input type="hidden" name="table" value="{e(table)}">
          <div class="controls">
            <select name="product" required>{options}</select>
            <input type="number" name="count" value="1" min="1" max="500">
            <button type="submit">Создать</button>
          </div>
        </form>
        """
    elif table == "keys":
        html += f"""
        <form method="post" action="{url_for('admin.add_record')}">
          <input type="hidden" name="bot" value="{e(bot_id)}">
          <input type="hidden" name="table" value="{e(table)}">
          <div class="controls">
            <input type="text" name="sku" required placeholder="SKU">
            <input type="text" name="license_key" required placeholder="Ключ">
            <button type="submit">Добавить</button>
          </div>
        </form>
        """
    elif table == "skus":
        html += f"""
        <form method="post" action="{url_for('admin.add_record')}">
          <input type="hidden" name="bot" value="{e(bot_id)}">
          <input type="hidden" name="table" value="{e(table)}">
          <div class="controls">
            <input type="text" name="sku" required placeholder="SKU">
            <input type="text" name="title" placeholder="Название">
            <button type="submit">Добавить</button>
          </div>
        </form>
        """
    elif table == "hidden_skus":
        html += f"""
        <form method="post" action="{url_for('admin.add_record')}">
          <input type="hidden" name="bot" value="{e(bot_id)}">
          <input type="hidden" name="table" value="{e(table)}">
          <div class="controls">
            <input type="text" name="sku" required placeholder="SKU">
            <button type="submit">Скрыть</button>
          </div>
        </form>
        """
    elif table == "sku_groups":
        html += f"""
        <form method="post" action="{url_for('admin.add_record')}">
          <input type="hidden" name="bot" value="{e(bot_id)}">
          <input type="hidden" name="table" value="{e(table)}">
          <div class="controls">
            <input type="text" name="sku" required placeholder="SKU">
            <select name="group_id" required>
              <option value="1">1 — Ключи</option>
              <option value="2">2 — Гифты</option>
            </select>
            <button type="submit">Сохранить</button>
          </div>
        </form>
        """
    else:
        html += "<p>Для этой таблицы форма не поддерживается.</p>"

    html += (
        f'<div style="margin-top:16px">'
        f'<a href="{url_for("admin.index", bot=bot_id, table=table)}">'
        f'<button type="button" class="gray">Назад</button></a></div></div>'
    )
    return render_page(html, current_bot=bot_id)


@admin.route("/add", methods=["POST"])
@login_required
def add_record():
    bot_id = request.form.get("bot", "").strip()
    table = request.form.get("table", "").strip()
    bot, table_meta = safe_get_bot_and_table(bot_id, table)
    if not bot or not table_meta:
        flash("Неверный bot или table", "error")
        return redirect(url_for("admin.index", bot=bot_id))

    ph = get_placeholder(bot_id)

    try:
        if table == "promo_codes":
            code = request.form.get("code", "").strip()
            product = request.form.get("product", "").strip()
            db_execute(bot_id, f"INSERT INTO promo_codes (code, product, status) VALUES ({ph}, {ph}, {ph})", (code, product, Status.FREE))
            flash("Промокод добавлен", "success")
        elif table == "accounts":
            data = request.form.get("data", "").strip()
            db_execute(bot_id, f"INSERT INTO accounts (data, status) VALUES ({ph}, {ph})", (data, Status.FREE))
            flash("Аккаунт добавлен", "success")
        elif table == "activation_links":
            product = request.form.get("product", "").strip()
            count = max(1, min(500, safe_int(request.form.get("count", "1"))))
            links = create_activation_links(bot_id, product, count)
            session["created_links"] = links
            flash(f"Создано ссылок: {count}", "success")
            return redirect(url_for("admin.activate_created_links", bot=bot_id))
        elif table == "keys":
            sku = request.form.get("sku", "").strip()
            license_key = request.form.get("license_key", "").strip()
            db_execute(bot_id, f"INSERT INTO keys (sku, license_key, status) VALUES ({ph}, {ph}, {ph})", (sku, license_key, Status.FREE))
            flash("Ключ добавлен", "success")
        elif table == "skus":
            sku = request.form.get("sku", "").strip()
            title = request.form.get("title", "").strip() or None
            _insert_or_ignore(bot_id, f"INSERT OR IGNORE INTO skus (sku, title) VALUES ({ph}, {ph})", f"INSERT INTO skus (sku, title) VALUES ({ph}, {ph}) ON CONFLICT DO NOTHING", (sku, title))
            if title is not None and BOTS[bot_id]["type"] == "sqlite":
                db_execute(bot_id, f"UPDATE skus SET title={ph} WHERE sku={ph}", (title, sku))
            flash("SKU сохранён", "success")
        elif table == "hidden_skus":
            sku = request.form.get("sku", "").strip()
            hide_sku(bot_id, sku)
            flash("SKU скрыт", "success")
        elif table == "sku_groups":
            sku = request.form.get("sku", "").strip()
            group_id = safe_int(request.form.get("group_id", "1"))
            set_sku_group(bot_id, sku, group_id)
            flash("Группа SKU сохранена", "success")
        else:
            flash("Для этой таблицы добавление не поддерживается", "error")
    except Exception as exc:
        logger.exception("Error adding record: %s", exc)
        flash(f"Ошибка добавления: {exc}", "error")

    return redirect(url_for("admin.index", bot=bot_id, table=table))


@admin.route("/form/bulk")
@login_required
def bulk_add_form():
    bot_id = request.args.get("bot", "").strip()
    table = request.args.get("table", "").strip()
    bot, table_meta = safe_get_bot_and_table(bot_id, table)
    if not bot or table not in ("promo_codes", "keys"):
        flash("Форма недоступна", "error")
        return redirect(url_for("admin.index", bot=bot_id))

    html = f'<div class="box"><h2>Массовая загрузка: {e(table_meta["name"])}</h2>'
    if table == "keys":
        html += f"""
        <form method="post" action="{url_for('admin.bulk_add')}">
          <input type="hidden" name="bot" value="{e(bot_id)}">
          <input type="hidden" name="table" value="{e(table)}">
          <div class="controls"><input type="text" name="sku" required placeholder="SKU"></div>
          <div style="margin-top:12px"><textarea name="items" required placeholder="Каждый ключ с новой строки"></textarea></div>
          <div style="margin-top:12px"><button type="submit" class="green">Загрузить</button></div>
        </form>
        """
    else:
        options = "".join([f'<option value="{e(k)}">{e(v)}</option>' for k, v in bot["products"].items() if k != "plus_account"])
        html += f"""
        <form method="post" action="{url_for('admin.bulk_add')}">
          <input type="hidden" name="bot" value="{e(bot_id)}">
          <input type="hidden" name="table" value="{e(table)}">
          <div class="controls"><select name="product">{options}</select></div>
          <div style="margin-top:12px"><textarea name="items" required placeholder="Каждый код с новой строки"></textarea></div>
          <div style="margin-top:12px"><button type="submit" class="green">Загрузить</button></div>
        </form>
        """
    html += (
        f'<div style="margin-top:16px">'
        f'<a href="{url_for("admin.index", bot=bot_id, table=table)}">'
        f'<button type="button" class="gray">Назад</button></a></div></div>'
    )
    return render_page(html, current_bot=bot_id)


@admin.route("/bulk-add", methods=["POST"])
@login_required
def bulk_add():
    bot_id = request.form.get("bot", "").strip()
    table = request.form.get("table", "").strip()

    if bot_id not in BOTS or table not in ("promo_codes", "keys"):
        flash("Неверные параметры", "error")
        return redirect(url_for("admin.index", bot=bot_id))

    items = [x.strip() for x in request.form.get("items", "").splitlines() if x.strip()]
    ph = get_placeholder(bot_id)

    if not items:
        flash("Нет данных для загрузки", "error")
        return redirect(url_for("admin.index", bot=bot_id, table=table))

    added = 0
    skipped = 0

    try:
        if table == "keys":
            sku = request.form.get("sku", "").strip()
            for key in items:
                try:
                    db_execute(bot_id, f"INSERT INTO keys (sku, license_key, status) VALUES ({ph}, {ph}, {ph})", (sku, key, Status.FREE))
                    added += 1
                except Exception:
                    skipped += 1
        elif table == "promo_codes":
            product = request.form.get("product", "").strip()
            for code in items:
                try:
                    _insert_or_ignore(bot_id, f"INSERT OR IGNORE INTO promo_codes (code, product, status) VALUES ({ph}, {ph}, {ph})", f"INSERT INTO promo_codes (code, product, status) VALUES ({ph}, {ph}, {ph}) ON CONFLICT DO NOTHING", (code, product, Status.FREE))
                    added += 1
                except Exception:
                    skipped += 1

        msg = f"Добавлено {added} из {len(items)}"
        if skipped:
            msg += f" (пропущено: {skipped})"
        flash(msg, "success")

    except Exception as exc:
        logger.exception("Bulk add error: %s", exc)
        flash(f"Ошибка: {exc}", "error")

    return redirect(url_for("admin.index", bot=bot_id, table=table))


@admin.route("/delete-post", methods=["POST"])
@login_required
def delete_record_post():
    bot_id = request.args.get("bot", "").strip()
    table = request.args.get("table", "").strip()
    record_id = request.form.get("id")

    if bot_id not in BOTS:
        flash("Неверный bot", "error")
        return redirect(url_for("admin.index"))

    bot, table_meta = safe_get_bot_and_table(bot_id, table)
    if not bot or not table_meta:
        flash("Неверная таблица", "error")
        return redirect(url_for("admin.index", bot=bot_id))

    if not table_meta.get("deletable"):
        flash("Удаление не поддерживается для этой таблицы", "error")
        return redirect(url_for("admin.index", bot=bot_id, table=table))

    ph = get_placeholder(bot_id)

    try:
        col = ID_COL_MAP.get(table)
        if not col:
            raise ValueError("Удаление не поддерживается")
        db_execute(bot_id, f"DELETE FROM {table} WHERE {col}={ph}", (record_id,))
        flash("Запись удалена", "success")
    except Exception as exc:
        logger.exception("Delete error: %s", exc)
        flash(f"Ошибка удаления: {exc}", "error")

    return redirect(url_for("admin.index", bot=bot_id, table=table))


# =========================
# ACTIVATE ROUTES
# =========================
@admin.route("/activate/create-links")
@login_required
def activate_create_links_form():
    bot_id = request.args.get("bot", "activate")
    if bot_id not in BOTS:
        bot_id = "activate"
    options = "".join([f'<option value="{e(k)}">{e(v)}</option>' for k, v in BOTS["activate"]["products"].items()])
    html = f"""
    <div class="box">
      <h2>Создать ссылки активации</h2>
      <form method="post" action="{url_for('admin.activate_create_links')}">
        <input type="hidden" name="bot" value="{e(bot_id)}">
        <div class="controls">
          <select name="product">{options}</select>
          <input type="number" name="count" value="10" min="1" max="500">
          <button type="submit" class="green">Создать</button>
        </div>
      </form>
    </div>
    """
    return render_page(html, current_bot=bot_id)


@admin.route("/activate/create-links", methods=["POST"])
@login_required
def activate_create_links():
    bot_id = request.form.get("bot", "activate")
    if bot_id not in BOTS:
        bot_id = "activate"
    product = request.form.get("product", "").strip()
    count = max(1, min(500, safe_int(request.form.get("count", "1"))))
    try:
        links = create_activation_links(bot_id, product, count)
        session["created_links"] = links
        return redirect(url_for("admin.activate_created_links", bot=bot_id))
    except Exception as exc:
        flash(f"Ошибка: {exc}", "error")
        return redirect(url_for("admin.index", bot=bot_id))


@admin.route("/activate/created-links")
@login_required
def activate_created_links():
    bot_id = request.args.get("bot", "activate")
    if bot_id not in BOTS:
        bot_id = "activate"
    links = session.pop("created_links", [])
    links_text = e(chr(10).join(links))
    html = f"""
    <div class="box">
      <h2>Созданные ссылки</h2>
      <textarea readonly style="width:100%;height:280px">{links_text}</textarea>
      <div style="margin-top:12px">
        <a href="{url_for('admin.index', bot=bot_id)}"><button type="button" class="gray">Назад</button></a>
      </div>
    </div>
    """
    return render_page(html, current_bot=bot_id)


@admin.route("/activate/upload")
@login_required
def activate_upload_form():
    bot_id = request.args.get("bot", "activate")
    if bot_id not in BOTS:
        bot_id = "activate"
    options = "".join([f'<option value="{e(k)}">{e(v)}</option>' for k, v in BOTS["activate"]["products"].items()])
    html = f"""
    <div class="box">
      <h2>Загрузить лицензии / аккаунты</h2>
      <form method="post" action="{url_for('admin.activate_upload_codes')}">
        <input type="hidden" name="bot" value="{e(bot_id)}">
        <div class="controls"><select name="product">{options}</select></div>
        <div style="margin-top:12px"><textarea name="codes" required placeholder="Каждый код / аккаунт с новой строки"></textarea></div>
        <div style="margin-top:12px"><button type="submit" class="green">Загрузить</button></div>
      </form>
    </div>
    """
    return render_page(html, current_bot=bot_id)


@admin.route("/activate/upload", methods=["POST"])
@login_required
def activate_upload_codes():
    bot_id = request.form.get("bot", "activate")
    if bot_id not in BOTS:
        bot_id = "activate"
    product = request.form.get("product", "").strip()
    codes = [x.strip() for x in request.form.get("codes", "").splitlines() if x.strip()]
    ph = get_placeholder(bot_id)
    added = 0
    skipped = 0

    try:
        if product == "plus_account":
            for c in codes:
                try:
                    db_execute(bot_id, f"INSERT INTO accounts (data, status) VALUES ({ph}, {ph})", (c, Status.FREE))
                    added += 1
                except Exception:
                    skipped += 1
        else:
            for c in codes:
                try:
                    _insert_or_ignore(bot_id, f"INSERT OR IGNORE INTO promo_codes (code, product, status) VALUES ({ph}, {ph}, {ph})", f"INSERT INTO promo_codes (code, product, status) VALUES ({ph}, {ph}, {ph}) ON CONFLICT DO NOTHING", (c, product, Status.FREE))
                    added += 1
                except Exception:
                    skipped += 1
        msg = f"Добавлено {added} из {len(codes)}"
        if skipped:
            msg += f" (пропущено: {skipped})"
        flash(msg, "success")
    except Exception as exc:
        flash(f"Ошибка: {exc}", "error")

    return redirect(url_for("admin.index", bot=bot_id))


@admin.route("/activate/check-link", methods=["GET", "POST"])
@login_required
def activate_check_link_form():
    bot_id = request.values.get("bot", "activate")
    if bot_id not in BOTS:
        bot_id = "activate"
    result_html = ""
    if request.method == "POST":
        token = parse_activate_token(request.form.get("token", ""))
        ph = get_placeholder(bot_id)
        row = db_fetch_one(bot_id, f"SELECT * FROM activation_links WHERE token={ph}", (token,))
        if not row:
            result_html = '<div class="flash error">Ссылка не найдена</div>'
        else:
            pname = BOTS[bot_id]["products"].get(row.get("product"), row.get("product"))
            result_html = f"""
            <div class="box">
              <p><strong>Token:</strong> <code>{e(str(row.get("token")))}</code></p>
              <p><strong>Product:</strong> {e(str(pname))}</p>
              <p><strong>Status:</strong> {badge_html(str(row.get("status") or ""))}</p>
              <p><strong>Attempts:</strong> {e(str(row.get("attempts", 0)))}/{MAX_ATTEMPTS}</p>
              <p><strong>CDK:</strong> <code>{e(str(row.get("cdk_code") or "-"))}</code></p>
              <p><strong>Error:</strong> <code>{e(str(row.get("last_error") or "-"))}</code></p>
              <p><strong>Created:</strong> {e(str(row.get("created_at") or "-"))}</p>
            </div>
            """
    html = f"""
    <div class="box">
      <h2>Проверить ссылку</h2>
      <form method="post">
        <input type="hidden" name="bot" value="{e(bot_id)}">
        <div class="controls">
          <input type="text" name="token" required placeholder="Вставь ссылку или токен">
          <button type="submit">Проверить</button>
        </div>
      </form>
    </div>
    {result_html}
    """
    return render_page(html, current_bot=bot_id)


@admin.route("/activate/find-order", methods=["GET", "POST"])
@login_required
def activate_find_order_form():
    bot_id = request.values.get("bot", "activate")
    if bot_id not in BOTS:
        bot_id = "activate"
    result_html = ""
    if request.method == "POST":
        order_id = request.form.get("order_id", "").strip()
        ph = get_placeholder(bot_id)
        row = db_fetch_one(bot_id, f"SELECT * FROM orders WHERE order_id={ph}", (order_id,))
        if not row:
            result_html = '<div class="flash error">Заказ не найден</div>'
        else:
            pname = BOTS[bot_id]["products"].get(row.get("product"), row.get("product"))
            token = row.get("activation_token")
            link = f"{PUBLIC_BASE_URL}/l/{e(str(token))}" if token else "нет"
            result_html = f"""
            <div class="box">
              <p><strong>ID:</strong> <code>{e(str(row.get("order_id")))}</code></p>
              <p><strong>Product:</strong> {e(str(pname))}</p>
              <p><strong>Status:</strong> {e(str(row.get("status")))}</p>
              <p><strong>Paid:</strong> {e(str(row.get("paid_at") or "-"))}</p>
              <p><strong>Link:</strong> {link}</p>
            </div>
            """
    html = f"""
    <div class="box">
      <h2>Найти заказ</h2>
      <form method="post">
        <input type="hidden" name="bot" value="{e(bot_id)}">
        <div class="controls">
          <input type="text" name="order_id" required placeholder="Order ID">
          <button type="submit">Найти</button>
        </div>
      </form>
    </div>
    {result_html}
    """
    return render_page(html, current_bot=bot_id)


@admin.route("/activate/keys")
@login_required
def activate_keys_choose():
    bot_id = request.args.get("bot", "activate")
    if bot_id not in BOTS:
        bot_id = "activate"
    options = "".join(
        [
            f'<a href="{url_for("admin.activate_keys_list", bot=bot_id, product=pid, page=1)}">'
            f'<button type="button">{e(name)}</button></a>'
            for pid, name in BOTS[bot_id]["products"].items()
        ]
    )
    html = f"""
    <div class="box">
      <h2>Лицензии по тарифу</h2>
      <div class="controls">{options}</div>
    </div>
    """
    return render_page(html, current_bot=bot_id)


@admin.route("/activate/keys/<product>")
@login_required
def activate_keys_list(product):
    bot_id = request.args.get("bot", "activate")
    if bot_id not in BOTS:
        bot_id = "activate"
    page = max(1, safe_int(request.args.get("page", 1)))
    rows, total, total_pages, page = get_activate_codes_page(bot_id, product, page)
    codes = [r["code"] for r in rows]
    pname = BOTS[bot_id]["products"].get(product, product)

    pager = '<div class="controls">'
    if page > 1:
        pager += f'<a href="{url_for("admin.activate_keys_list", bot=bot_id, product=product, page=page - 1)}"><button type="button" class="gray">← Назад</button></a>'
    pager += f"<span>Страница {page}/{total_pages} | Всего: {total}</span>"
    if page < total_pages:
        pager += f'<a href="{url_for("admin.activate_keys_list", bot=bot_id, product=product, page=page + 1)}"><button type="button" class="gray">Вперёд →</button></a>'
    pager += "</div>"

    codes_text = e(chr(10).join(codes)) if codes else "Пусто"

    html = f"""
    <div class="box">
      <h2>{e(str(pname))}</h2>
      {pager}
      <pre style="margin-top:12px">{codes_text}</pre>
      <div style="margin-top:12px">
        <a href="{url_for('admin.activate_keys_choose', bot=bot_id)}"><button type="button" class="gray">Назад к тарифам</button></a>
      </div>
    </div>
    """
    return render_page(html, current_bot=bot_id)


@admin.route("/activate/links-by-product")
@login_required
def activate_links_choose():
    bot_id = request.args.get("bot", "activate")
    if bot_id not in BOTS:
        bot_id = "activate"
    options = "".join(
        [
            f'<a href="{url_for("admin.activate_links_list", bot=bot_id, product=pid, page=1)}">'
            f'<button type="button">{e(name)}</button></a>'
            for pid, name in BOTS[bot_id]["products"].items()
        ]
    )
    html = f"""
    <div class="box">
      <h2>Ссылки по тарифу</h2>
      <div class="controls">{options}</div>
    </div>
    """
    return render_page(html, current_bot=bot_id)


@admin.route("/activate/links-by-product/<product>")
@login_required
def activate_links_list(product):
    bot_id = request.args.get("bot", "activate")
    if bot_id not in BOTS:
        bot_id = "activate"
    page = max(1, safe_int(request.args.get("page", 1)))
    ph = get_placeholder(bot_id)

    total = db_fetch_val(bot_id, f"SELECT COUNT(*) FROM activation_links WHERE product={ph}", (product,))
    total_pages = max(1, math.ceil(total / PAGE_SIZE_CODES)) if total else 1
    page = max(1, min(page, total_pages))
    offset = (page - 1) * PAGE_SIZE_CODES

    if BOTS[bot_id]["type"] == "postgresql":
        rows = db_fetch_all(bot_id, f"SELECT token, status, created_at FROM activation_links WHERE product={ph} ORDER BY created_at DESC LIMIT {ph} OFFSET {ph}", (product, PAGE_SIZE_CODES, offset))
    else:
        rows = db_fetch_all(bot_id, f"SELECT token, status, created_at FROM activation_links WHERE product={ph} ORDER BY created_at DESC LIMIT {PAGE_SIZE_CODES} OFFSET {offset}", (product,))

    pname = BOTS[bot_id]["products"].get(product, product)
    base_url = BOTS[bot_id].get("public_base_url", PUBLIC_BASE_URL)
    prefix = "/a/" if product == "plus_account" else "/l/"
    links = [f'{base_url}{prefix}{r["token"]}' for r in rows]

    pager = '<div class="controls">'
    if page > 1:
        pager += f'<a href="{url_for("admin.activate_links_list", bot=bot_id, product=product, page=page - 1)}"><button type="button" class="gray">← Назад</button></a>'
    pager += f"<span>Страница {page}/{total_pages} | Всего: {total}</span>"
    if page < total_pages:
        pager += f'<a href="{url_for("admin.activate_links_list", bot=bot_id, product=product, page=page + 1)}"><button type="button" class="gray">Вперёд →</button></a>'
    pager += "</div>"

    links_text = e(chr(10).join(links)) if links else "Пусто"

    html = f"""
    <div class="box">
      <h2>Ссылки: {e(str(pname))}</h2>
      {pager}
      <pre style="margin-top:12px">{links_text}</pre>
      <div style="margin-top:12px">
        <a href="{url_for('admin.activate_links_choose', bot=bot_id)}"><button type="button" class="gray">Назад к тарифам</button></a>
      </div>
    </div>
    """
    return render_page(html, current_bot=bot_id)


# =========================
# MARKET ROUTES
# =========================
@admin.route("/market/add-sku", methods=["GET", "POST"])
@login_required
def market_add_sku_form():
    bot_id = request.values.get("bot", "")
    if bot_id not in BOTS:
        flash("Неверный bot", "error")
        return redirect(url_for("admin.index"))

    if request.method == "POST":
        sku = request.form.get("sku", "").strip()
        title = request.form.get("title", "").strip() or None
        try:
            ph = get_placeholder(bot_id)
            _insert_or_ignore(bot_id, f"INSERT OR IGNORE INTO skus (sku, title) VALUES ({ph}, {ph})", f"INSERT INTO skus (sku, title) VALUES ({ph}, {ph}) ON CONFLICT DO NOTHING", (sku, title))
            if title is not None and BOTS[bot_id]["type"] == "sqlite":
                db_execute(bot_id, f"UPDATE skus SET title={ph} WHERE sku={ph}", (title, sku))
            flash("SKU сохранён", "success")
            return redirect(url_for("admin.index", bot=bot_id))
        except Exception as exc:
            flash(f"Ошибка: {exc}", "error")

    html = f"""
    <div class="box">
      <h2>Добавить SKU</h2>
      <form method="post">
        <input type="hidden" name="bot" value="{e(bot_id)}">
        <div class="controls">
          <input type="text" name="sku" required placeholder="MRKT-XXXXX">
          <input type="text" name="title" placeholder="Название">
          <button type="submit" class="green">Сохранить</button>
        </div>
      </form>
    </div>
    """
    return render_page(html, current_bot=bot_id)


@admin.route("/market/upload-keys", methods=["GET", "POST"])
@login_required
def market_upload_keys_form():
    bot_id = request.values.get("bot", "")
    if bot_id not in BOTS:
        flash("Неверный bot", "error")
        return redirect(url_for("admin.index"))

    if request.method == "POST":
        sku = request.form.get("sku", "").strip()
        keys = [x.strip() for x in request.form.get("keys", "").splitlines() if x.strip()]
        ph = get_placeholder(bot_id)
        added = 0
        skipped = 0
        try:
            for k in keys:
                try:
                    db_execute(bot_id, f"INSERT INTO keys (sku, license_key, status) VALUES ({ph}, {ph}, {ph})", (sku, k, Status.FREE))
                    added += 1
                except Exception:
                    skipped += 1
            msg = f"Добавлено {added} ключей"
            if skipped:
                msg += f" (пропущено: {skipped})"
            flash(msg, "success")
            return redirect(url_for("admin.index", bot=bot_id))
        except Exception as exc:
            flash(f"Ошибка: {exc}", "error")

    skus = list_visible_skus(bot_id)
    opts = "".join([f'<option value="{e(s["sku"])}">{(e(s["title"]) + " (" + e(s["sku"]) + ")") if s.get("title") else e(s["sku"])}</option>' for s in skus])
    html = f"""
    <div class="box">
      <h2>Загрузить ключи</h2>
      <form method="post">
        <input type="hidden" name="bot" value="{e(bot_id)}">
        <div class="controls"><select name="sku">{opts}</select></div>
        <div style="margin-top:12px"><textarea name="keys" required placeholder="Каждый ключ с новой строки"></textarea></div>
        <div style="margin-top:12px"><button type="submit" class="green">Загрузить</button></div>
      </form>
    </div>
    """
    return render_page(html, current_bot=bot_id)


@admin.route("/market/find-order-keys", methods=["GET", "POST"])
@login_required
def market_find_order_keys_form():
    bot_id = request.values.get("bot", "")
    if bot_id not in BOTS:
        flash("Неверный bot", "error")
        return redirect(url_for("admin.index"))

    result_html = ""
    if request.method == "POST":
        order_id = request.form.get("order_id", "").strip()
        ph = get_placeholder(bot_id)
        rows = db_fetch_all(bot_id, f"SELECT sku, license_key, status, used_at FROM keys WHERE order_id={ph} ORDER BY used_at, id", (order_id,))
        if not rows:
            result_html = '<div class="box"><p>Ключи не найдены</p></div>'
        else:
            parts = []
            for r in rows:
                parts.append(f"<p><strong>SKU:</strong> {e(str(r['sku']))}<br><strong>Ключ:</strong> <code>{e(str(r['license_key']))}</code><br><strong>Статус:</strong> {e(str(r['status']))}<br><strong>Выдан:</strong> {e(str(r.get('used_at') or '-'))}</p>")
            result_html = f'<div class="box"><h3>Найдено: {len(rows)}</h3>{"".join(parts)}</div>'

    html = f"""
    <div class="box">
      <h2>Ключи по заказу</h2>
      <form method="post">
        <input type="hidden" name="bot" value="{e(bot_id)}">
        <div class="controls">
          <input type="text" name="order_id" required placeholder="Order ID">
          <button type="submit">Найти</button>
        </div>
      </form>
    </div>
    {result_html}
    """
    return render_page(html, current_bot=bot_id)


@admin.route("/market/hide-show", methods=["GET", "POST"])
@login_required
def market_hide_show_form():
    bot_id = request.values.get("bot", "")
    if bot_id not in BOTS:
        flash("Неверный bot", "error")
        return redirect(url_for("admin.index"))

    if request.method == "POST":
        sku = request.form.get("sku", "").strip()
        action = request.form.get("action", "").strip()
        try:
            if action == "hide":
                hide_sku(bot_id, sku)
                flash(f"SKU скрыт: {sku}", "success")
            else:
                unhide_sku(bot_id, sku)
                flash(f"SKU возвращён: {sku}", "success")
            return redirect(url_for("admin.index", bot=bot_id))
        except Exception as exc:
            flash(f"Ошибка: {exc}", "error")

    html = f"""
    <div class="box">
      <h2>Скрыть / показать SKU</h2>
      <form method="post">
        <input type="hidden" name="bot" value="{e(bot_id)}">
        <div class="controls">
          <input type="text" name="sku" required placeholder="MRKT-XXXXX">
          <select name="action">
            <option value="hide">Скрыть</option>
            <option value="show">Показать</option>
          </select>
          <button type="submit">Применить</button>
        </div>
      </form>
    </div>
    """
    return render_page(html, current_bot=bot_id)


@admin.route("/market/groups", methods=["GET", "POST"])
@login_required
def market_groups_form():
    bot_id = request.values.get("bot", "")
    if bot_id not in BOTS:
        flash("Неверный bot", "error")
        return redirect(url_for("admin.index"))

    if request.method == "POST":
        sku = request.form.get("sku", "").strip()
        action = request.form.get("group_action", "").strip()
        try:
            if action == "remove":
                remove_sku_group(bot_id, sku)
                flash("SKU удалён из группы", "success")
            else:
                group_id = safe_int(request.form.get("group_id", "1"))
                set_sku_group(bot_id, sku, group_id)
                flash("SKU привязан к группе", "success")
            return redirect(url_for("admin.index", bot=bot_id))
        except Exception as exc:
            flash(f"Ошибка: {exc}", "error")

    rows = db_fetch_all(bot_id, "SELECT sku, group_id FROM sku_groups ORDER BY sku")
    tr = ("".join([f"<tr><td>{e(str(r['sku']))}</td><td>{e(str(r['group_id']))}</td></tr>" for r in rows]) or "<tr><td colspan='2'>Пусто</td></tr>")
    html = f"""
    <div class="grid2">
      <div class="box">
        <h2>Группы SKU</h2>
        <form method="post">
          <input type="hidden" name="bot" value="{e(bot_id)}">
          <div class="controls">
            <input type="text" name="sku" required placeholder="SKU">
            <select name="group_action">
              <option value="set">Привязать</option>
              <option value="remove">Удалить из группы</option>
            </select>
            <select name="group_id">
              <option value="1">1 — Ключи</option>
              <option value="2">2 — Гифты</option>
            </select>
            <button type="submit">Сохранить</button>
          </div>
        </form>
      </div>
      <div class="box">
        <h3>Текущие привязки</h3>
        <table><thead><tr><th>SKU</th><th>Группа</th></tr></thead><tbody>{tr}</tbody></table>
      </div>
    </div>
    """
    return render_page(html, current_bot=bot_id)


@admin.route("/market/sku-tools", methods=["GET", "POST"])
@login_required
def market_sku_tools():
    bot_id = request.values.get("bot", "")
    if bot_id not in BOTS:
        flash("Неверный bot", "error")
        return redirect(url_for("admin.index"))

    sku = request.values.get("sku", "").strip()

    if request.method == "POST":
        action = request.form.get("tool_action", "").strip()
        sku = request.form.get("sku", "").strip()
        try:
            if action == "title":
                title = request.form.get("title", "").strip() or None
                set_sku_title(bot_id, sku, title)
                flash("Название сохранено", "success")
            elif action == "slip":
                slip = request.form.get("slip_text", "")
                if slip.strip().lower() in ("стандарт", "default", "reset"):
                    slip = None
                set_sku_slip_text(bot_id, sku, slip)
                flash("Текст отправки сохранён", "success")
            return redirect(url_for("admin.market_sku_tools", bot=bot_id, sku=sku))
        except Exception as exc:
            flash(f"Ошибка: {exc}", "error")

    title = get_sku_title(bot_id, sku) if sku else ""
    slip = get_sku_slip_text(bot_id, sku) if sku else ""

    html = f"""
    <div class="box">
      <h2>Инструменты SKU</h2>
      <form method="get" class="controls">
        <input type="hidden" name="bot" value="{e(bot_id)}">
        <input type="text" name="sku" value="{e(sku)}" placeholder="SKU">
        <button type="submit">Открыть</button>
      </form>
    </div>
    """

    if sku:
        html += f"""
        <div class="grid2">
          <div class="box">
            <h3>Подпись товара</h3>
            <form method="post">
              <input type="hidden" name="bot" value="{e(bot_id)}">
              <input type="hidden" name="sku" value="{e(sku)}">
              <input type="hidden" name="tool_action" value="title">
              <input type="text" name="title" value="{e(title or '')}" placeholder="Название">
              <div style="margin-top:12px"><button type="submit">Сохранить</button></div>
            </form>
          </div>
          <div class="box">
            <h3>Текст отправки</h3>
            <form method="post">
              <input type="hidden" name="bot" value="{e(bot_id)}">
              <input type="hidden" name="sku" value="{e(sku)}">
              <input type="hidden" name="tool_action" value="slip">
              <textarea name="slip_text">{e(slip or '')}</textarea>
              <div class="small">Чтобы вернуть стандартный текст, введи: стандарт</div>
              <div style="margin-top:12px"><button type="submit">Сохранить</button></div>
            </form>
          </div>
        </div>
        <div class="box">
          <a href="{url_for('admin.market_free_keys_by_sku', bot=bot_id, sku=sku, page=1)}"><button type="button" class="gray">Свободные ключи этого SKU</button></a>
        </div>
        """
    return render_page(html, current_bot=bot_id)


@admin.route("/market/free-keys")
@login_required
def market_free_keys_by_sku():
    bot_id = request.args.get("bot", "")
    if bot_id not in BOTS:
        flash("Неверный bot", "error")
        return redirect(url_for("admin.index"))

    sku = request.args.get("sku", "").strip()
    page = max(1, safe_int(request.args.get("page", 1)))
    rows, total, total_pages, page = get_free_keys_page_for_sku(bot_id, sku, page, 30)

    label = get_sku_title(bot_id, sku)
    title = f"{label} ({sku})" if label else sku

    pager = '<div class="controls">'
    if page > 1:
        pager += f'<a href="{url_for("admin.market_free_keys_by_sku", bot=bot_id, sku=sku, page=page - 1)}"><button type="button" class="gray">← Назад</button></a>'
    pager += f"<span>Страница {page}/{total_pages} | Всего: {total}</span>"
    if page < total_pages:
        pager += f'<a href="{url_for("admin.market_free_keys_by_sku", bot=bot_id, sku=sku, page=page + 1)}"><button type="button" class="gray">Вперёд →</button></a>'
    pager += "</div>"

    lines = [r["license_key"] for r in rows]
    lines_text = e(chr(10).join(lines)) if lines else "Пусто"

    html = f"""
    <div class="box">
      <h2>Свободные ключи: {e(str(title))}</h2>
      {pager}
      <pre style="margin-top:12px">{lines_text}</pre>
      <div style="margin-top:12px">
        <a href="{url_for('admin.market_sku_tools', bot=bot_id, sku=sku)}"><button type="button" class="gray">Назад</button></a>
      </div>
    </div>
    """
    return render_page(html, current_bot=bot_id)


@admin.route("/market/manual-issue", methods=["GET", "POST"])
@login_required
def market_manual_issue_form():
    bot_id = request.values.get("bot", "")
    if bot_id not in BOTS:
        flash("Неверный bot", "error")
        return redirect(url_for("admin.index"))

    result_html = ""

    if request.method == "POST":
        order_id = request.form.get("order_id", "").strip()
        manual_text = request.form.get("manual_codes", "").strip()

        if not market_api_ready():
            flash("Не заданы переменные Yandex Market API", "error")
            return redirect(url_for("admin.market_manual_issue_form", bot=bot_id))

        try:
            order = fetch_order_from_api(order_id)
            if not order:
                flash("Заказ не найден в API Маркета", "error")
                return redirect(url_for("admin.market_manual_issue_form", bot=bot_id))

            if is_order_processed(bot_id, order_id):
                flash("Внимание: заказ уже помечен как обработанный", "info")

            items = order.get("items", [])
            if not items:
                flash("В заказе нет items", "error")
                return redirect(url_for("admin.market_manual_issue_form", bot=bot_id))

            manual_map = {}
            if manual_text:
                for line in manual_text.splitlines():
                    line = line.strip()
                    if not line or "|" not in line:
                        continue
                    item_id, key = line.split("|", 1)
                    manual_map.setdefault(str(item_id).strip(), []).append(key.strip())

            all_payload = []
            saved_manual = []

            for it in items:
                item_id = str(it.get("id"))
                sku = it.get("offerId") or it.get("shopSku")
                count = safe_int(it.get("count", 1))

                pool_keys = take_free_keys_for_sku(bot_id, sku, order_id, count)
                final_keys = list(pool_keys)

                if len(final_keys) < count:
                    need = count - len(final_keys)
                    extra = manual_map.get(item_id, [])[:need]
                    final_keys.extend(extra)
                    if extra:
                        saved_manual.append((sku, item_id, extra))

                if not final_keys:
                    continue

                slip_text = get_sku_slip_text(bot_id, sku) or SLIP_TEXT_DEFAULT
                all_payload.append({"id": int(item_id), "codes": final_keys, "slip": slip_text, "activate_till": ACTIVATE_TILL})

            if not all_payload:
                flash("Не удалось собрать коды для отправки", "error")
                return redirect(url_for("admin.market_manual_issue_form", bot=bot_id))

            resp = deliver_digital_goods_manual(order_id, all_payload)
            result_html = f"""
            <div class="box">
              <h3>Ответ Маркета</h3>
              <p><strong>HTTP:</strong> {e(str(resp.status_code))}</p>
              <pre>{e(resp.text[:3000])}</pre>
            </div>
            """

            if resp.status_code == 200:
                for sku_val, _item_id, keys in saved_manual:
                    save_used_keys(bot_id, sku_val, order_id, keys)
                mark_order_processed(bot_id, order_id)
                flash("Ручная выдача завершена", "success")
            else:
                flash("Маркет вернул не 200", "error")

        except Exception as exc:
            logger.exception("Manual issue error: %s", exc)
            flash(f"Ошибка ручной выдачи: {exc}", "error")

    html = f"""
    <div class="box">
      <h2>Ручная выдача</h2>
      <p class="small">Формат ручных ключей: <code>item_id|КЛЮЧ</code> — каждый ключ на новой строке.</p>
      <form method="post">
        <input type="hidden" name="bot" value="{e(bot_id)}">
        <div class="controls">
          <input type="text" name="order_id" required placeholder="Order ID">
        </div>
        <div style="margin-top:12px">
          <textarea name="manual_codes" placeholder="123456789|KEY-ONE&#10;123456789|KEY-TWO"></textarea>
        </div>
        <div style="margin-top:12px">
          <button type="submit" class="red">Выполнить ручную выдачу</button>
        </div>
      </form>
    </div>
    {result_html}
    """
    return render_page(html, current_bot=bot_id)


# =========================
# CABINET MANAGEMENT ROUTES
# =========================

@admin.route("/activate/withdrawals", methods=["GET", "POST"])
@login_required
def activate_withdrawals():
    bot_id = request.values.get("bot", "activate")
    if bot_id not in BOTS:
        bot_id = "activate"
    ph = get_placeholder(bot_id)

    if request.method == "POST":
        action = request.form.get("action", "").strip()
        wid = safe_int(request.form.get("withdraw_id", 0), 0)
        current_status = request.form.get("current_status", "new")

        if not wid:
            flash("Неверный ID заявки", "error")
            return redirect(url_for("admin.activate_withdrawals", bot=bot_id, status=current_status))

        row = db_fetch_one(bot_id, f"SELECT * FROM withdraw_requests WHERE id={ph}", (wid,))
        if not row:
            flash("Заявка не найдена", "error")
            return redirect(url_for("admin.activate_withdrawals", bot=bot_id, status=current_status))

        if row["status"] != "new":
            flash("Заявка уже обработана", "error")
            return redirect(url_for("admin.activate_withdrawals", bot=bot_id, status=current_status))

        try:
            if action == "approve":
                with get_db(bot_id) as conn:
                    cur = conn.cursor()
                    cur.execute(f"UPDATE withdraw_requests SET status='approved', processed_at=NOW() WHERE id={ph} AND status='new'", (wid,))
                    cur.execute(f"INSERT INTO wallet_transactions(user_id, tx_type, amount, meta_json) VALUES({ph}, 'withdraw_done', {ph}, {ph})", (row["user_id"], -row["amount"], f'{{"withdraw_request_id": {wid}}}'))
                    conn.commit()
                flash(f"Заявка #{wid} одобрена", "success")

            elif action == "reject":
                with get_db(bot_id) as conn:
                    cur = conn.cursor()
                    cur.execute(f"UPDATE withdraw_requests SET status='rejected', processed_at=NOW() WHERE id={ph} AND status='new'", (wid,))
                    cur.execute(f"UPDATE users SET balance = balance + {ph} WHERE id={ph}", (row["amount"], row["user_id"]))
                    cur.execute(f"INSERT INTO wallet_transactions(user_id, tx_type, amount, meta_json) VALUES({ph}, 'withdraw_rejected', {ph}, {ph})", (row["user_id"], row["amount"], f'{{"withdraw_request_id": {wid}}}'))
                    conn.commit()
                flash(f"Заявка #{wid} отклонена, баланс возвращён", "success")

        except Exception as exc:
            logger.exception("Withdrawal action error: %s", exc)
            flash(f"Ошибка: {exc}", "error")

        return redirect(url_for("admin.activate_withdrawals", bot=bot_id, status=current_status))

    # GET
    status_filter = request.args.get("status", "new").strip()
    page = max(1, safe_int(request.args.get("page", 1)))

    where_parts = []
    params = []
    if status_filter in ("new", "approved", "rejected"):
        where_parts.append(f"w.status={ph}")
        params.append(status_filter)

    where_clause = ("WHERE " + " AND ".join(where_parts)) if where_parts else ""

    total = db_fetch_val(bot_id, f"SELECT COUNT(*) FROM withdraw_requests w {where_clause}", tuple(params))
    total_pages = max(1, math.ceil(total / PAGE_SIZE)) if total else 1
    page = max(1, min(page, total_pages))
    offset = (page - 1) * PAGE_SIZE

    rows = db_fetch_all(
        bot_id,
        f"SELECT w.id, w.user_id, w.amount, w.status, w.note, w.created_at, w.processed_at, u.email "
        f"FROM withdraw_requests w JOIN users u ON u.id = w.user_id "
        f"{where_clause} ORDER BY w.created_at DESC LIMIT {ph} OFFSET {ph}",
        tuple(params) + (PAGE_SIZE, offset),
    )

    filter_bar = '<div class="filter-bar"><span class="label">Статус:</span>'
    for st_val, st_label in [("", "Все"), ("new", "Новые"), ("approved", "Одобренные"), ("rejected", "Отклонённые")]:
        is_active = "active" if status_filter == st_val else ""
        if st_val == "" and status_filter not in ("new", "approved", "rejected"):
            is_active = "active"
        filter_bar += f'<a class="{is_active}" href="{url_for("admin.activate_withdrawals", bot=bot_id, status=st_val)}">{e(st_label)}</a>'
    filter_bar += "</div>"

    if rows:
        body_rows = []
        for r in rows:
            actions_html = ""
            if r["status"] == "new":
                actions_html = (
                    f'<form method="post" style="display:inline">'
                    f'<input type="hidden" name="bot" value="{e(bot_id)}">'
                    f'<input type="hidden" name="withdraw_id" value="{e(str(r["id"]))}">'
                    f'<input type="hidden" name="current_status" value="{e(status_filter)}">'
                    f'<input type="hidden" name="action" value="approve">'
                    f'<button type="submit" class="green" style="margin-right:4px">✅ Одобрить</button>'
                    f'</form>'
                    f'<form method="post" style="display:inline" onsubmit="return confirm(\'Отклонить заявку #{r["id"]}? Баланс {r["amount"]} ₽ будет возвращён.\')">'
                    f'<input type="hidden" name="bot" value="{e(bot_id)}">'
                    f'<input type="hidden" name="withdraw_id" value="{e(str(r["id"]))}">'
                    f'<input type="hidden" name="current_status" value="{e(status_filter)}">'
                    f'<input type="hidden" name="action" value="reject">'
                    f'<button type="submit" class="red">❌ Отклонить</button>'
                    f'</form>'
                )

            note_val = r.get("note") or "—"
            note_short = note_val[:80] + ("..." if len(note_val) > 80 else "")

            body_rows.append(
                f"<tr>"
                f"<td><strong>#{e(str(r['id']))}</strong></td>"
                f'<td><a href="{url_for("admin.activate_cabinet_user", user_id=r["user_id"], bot=bot_id)}">{e(str(r["email"]))}</a></td>'
                f"<td><strong>{e(str(r['amount']))} ₽</strong></td>"
                f"<td>{badge_html(str(r['status']))}</td>"
                f'<td><code title="{e(note_val)}">{e(note_short)}</code></td>'
                f'<td class="small">{e(str(r["created_at"] or ""))}</td>'
                f'<td class="small">{e(str(r["processed_at"] or "—"))}</td>'
                f"<td>{actions_html}</td>"
                f"</tr>"
            )

        pager = '<div class="controls" style="margin-top:12px">'
        if page > 1:
            pager += f'<a href="{url_for("admin.activate_withdrawals", bot=bot_id, status=status_filter, page=page - 1)}"><button type="button" class="gray">← Назад</button></a>'
        pager += f"<span>Страница {page} из {total_pages} | Всего: {total}</span>"
        if page < total_pages:
            pager += f'<a href="{url_for("admin.activate_withdrawals", bot=bot_id, status=status_filter, page=page + 1)}"><button type="button" class="gray">Вперёд →</button></a>'
        pager += "</div>"

        table_html = (
            f'<table><thead><tr><th>ID</th><th>Email</th><th>Сумма</th><th>Статус</th><th>Реквизиты</th><th>Создана</th><th>Обработана</th><th>Действия</th></tr></thead>'
            f'<tbody>{"".join(body_rows)}</tbody></table>{pager}'
        )
    else:
        table_html = "<p>Нет заявок</p>"

    html = (
        f'<div class="box"><h2>💰 Заявки на вывод</h2>{filter_bar}'
        f'<div style="margin-top:12px"><a href="{url_for("admin.index", bot=bot_id)}"><button type="button" class="gray">← Назад</button></a></div></div>'
        f'<div class="box">{table_html}</div>'
    )
    return render_page(html, current_bot=bot_id)


@admin.route("/activate/cabinet-users")
@login_required
def activate_cabinet_users():
    bot_id = request.args.get("bot", "activate")
    if bot_id not in BOTS:
        bot_id = "activate"
    ph = get_placeholder(bot_id)
    page = max(1, safe_int(request.args.get("page", 1)))
    q = request.args.get("q", "").strip()

    where_clause = ""
    params = []
    if q:
        where_clause = f"WHERE email ILIKE {ph}"
        params.append(f"%{q}%")

    total = db_fetch_val(bot_id, f"SELECT COUNT(*) FROM users {where_clause}", tuple(params))
    total_pages = max(1, math.ceil(total / PAGE_SIZE)) if total else 1
    page = max(1, min(page, total_pages))
    offset = (page - 1) * PAGE_SIZE

    rows = db_fetch_all(bot_id, f"SELECT id, email, balance, ref_code, referrer_id, created_at, is_active FROM users {where_clause} ORDER BY created_at DESC LIMIT {ph} OFFSET {ph}", tuple(params) + (PAGE_SIZE, offset))

    search_form = (
        f'<form method="get" class="controls">'
        f'<input type="hidden" name="bot" value="{e(bot_id)}">'
        f'<input type="text" name="q" value="{e(q)}" placeholder="Поиск по email">'
        f'<button type="submit">Найти</button>'
        f'<a href="{url_for("admin.activate_cabinet_users", bot=bot_id)}"><button type="button" class="gray">Сбросить</button></a>'
        f'<a href="{url_for("admin.index", bot=bot_id)}"><button type="button" class="gray">← Назад</button></a>'
        f'</form>'
    )

    if rows:
        body_rows = []
        for r in rows:
            active_badge = badge_html("active") if r["is_active"] else badge_html("blocked")
            body_rows.append(
                f"<tr>"
                f"<td>{e(str(r['id']))}</td>"
                f'<td><a href="{url_for("admin.activate_cabinet_user", user_id=r["id"], bot=bot_id)}">{e(str(r["email"]))}</a></td>'
                f"<td><strong>{e(str(r['balance']))} ₽</strong></td>"
                f"<td><code>{e(str(r['ref_code']))}</code></td>"
                f"<td>{e(str(r['referrer_id'] or '—'))}</td>"
                f'<td class="small">{e(str(r["created_at"] or ""))}</td>'
                f"<td>{active_badge}</td>"
                f'<td><a href="{url_for("admin.activate_cabinet_user", user_id=r["id"], bot=bot_id)}"><button type="button" class="gray">Подробнее</button></a></td>'
                f"</tr>"
            )

        pager = '<div class="controls" style="margin-top:12px">'
        if page > 1:
            pager += f'<a href="{url_for("admin.activate_cabinet_users", bot=bot_id, q=q, page=page - 1)}"><button type="button" class="gray">← Назад</button></a>'
        pager += f"<span>Страница {page} из {total_pages} | Всего: {total}</span>"
        if page < total_pages:
            pager += f'<a href="{url_for("admin.activate_cabinet_users", bot=bot_id, q=q, page=page + 1)}"><button type="button" class="gray">Вперёд →</button></a>'
        pager += "</div>"

        table_html = (
            f'<table><thead><tr><th>ID</th><th>Email</th><th>Баланс</th><th>Реф. код</th><th>Реферер</th><th>Регистрация</th><th>Статус</th><th></th></tr></thead>'
            f'<tbody>{"".join(body_rows)}</tbody></table>{pager}'
        )
    else:
        table_html = "<p>Пользователи не найдены</p>"

    html = f'<div class="box"><h2>👥 Пользователи кабинета</h2>{search_form}</div><div class="box">{table_html}</div>'
    return render_page(html, current_bot=bot_id)


@admin.route("/activate/cabinet-user/<int:user_id>", methods=["GET", "POST"])
@login_required
def activate_cabinet_user(user_id):
    bot_id = request.values.get("bot", "activate")
    if bot_id not in BOTS:
        bot_id = "activate"
    ph = get_placeholder(bot_id)
    products_dict = BOTS[bot_id].get("products", {})

    if request.method == "POST":
        action = request.form.get("action", "").strip()
        try:
            if action == "adjust_balance":
                amount = safe_int(request.form.get("amount", 0), 0)
                note = request.form.get("note", "").strip() or "Ручная корректировка"
                if amount == 0:
                    flash("Сумма не может быть 0", "error")
                else:
                    with get_db(bot_id) as conn:
                        cur = conn.cursor()
                        cur.execute(f"UPDATE users SET balance = balance + {ph} WHERE id={ph}", (amount, user_id))
                        meta = f'{{"admin_note": "{note}"}}'
                        tx_type = "bonus" if amount > 0 else "correction"
                        cur.execute(f"INSERT INTO wallet_transactions(user_id, tx_type, amount, meta_json) VALUES({ph}, {ph}, {ph}, {ph})", (user_id, tx_type, amount, meta))
                        conn.commit()
                    sign = "+" if amount > 0 else ""
                    flash(f"Баланс скорректирован: {sign}{amount} ₽", "success")

            elif action == "kill_sessions":
                db_execute(bot_id, f"DELETE FROM user_sessions WHERE user_id={ph}", (user_id,))
                flash("Все сессии пользователя завершены", "success")

            elif action == "reset_password":
                new_password = secrets.token_urlsafe(12)
                pw_hash = hash_password(new_password)
                db_execute(bot_id, f"UPDATE users SET password_hash={ph} WHERE id={ph}", (pw_hash, user_id))
                db_execute(bot_id, f"DELETE FROM user_sessions WHERE user_id={ph}", (user_id,))
                flash(f"Новый пароль: {new_password} — скопируйте и передайте пользователю!", "info")

        except Exception as exc:
            logger.exception("Cabinet user action error: %s", exc)
            flash(f"Ошибка: {exc}", "error")

        return redirect(url_for("admin.activate_cabinet_user", user_id=user_id, bot=bot_id))

    user = db_fetch_one(bot_id, f"SELECT * FROM users WHERE id={ph}", (user_id,))
    if not user:
        flash("Пользователь не найден", "error")
        return redirect(url_for("admin.activate_cabinet_users", bot=bot_id))

    referrer_email = "—"
    if user.get("referrer_id"):
        ref_row = db_fetch_one(bot_id, f"SELECT email FROM users WHERE id={ph}", (user["referrer_id"],))
        if ref_row:
            referrer_email = ref_row["email"]

    orders = db_fetch_all(bot_id, f"SELECT order_id, product, amount, status, created_at, paid_at FROM orders WHERE user_id={ph} ORDER BY created_at DESC LIMIT 50", (user_id,))
    referrals = db_fetch_all(bot_id, f"SELECT u.email, r.total_earned, r.created_at, r.status FROM referrals r JOIN users u ON u.id = r.referral_id WHERE r.referrer_id={ph} ORDER BY r.created_at DESC LIMIT 50", (user_id,))
    transactions = db_fetch_all(bot_id, f"SELECT tx_type, amount, meta_json, created_at FROM wallet_transactions WHERE user_id={ph} ORDER BY created_at DESC LIMIT 50", (user_id,))
    withdrawals = db_fetch_all(bot_id, f"SELECT id, amount, status, note, created_at, processed_at FROM withdraw_requests WHERE user_id={ph} ORDER BY created_at DESC LIMIT 50", (user_id,))

    profile_html = (
        f'<div class="box"><h2>👤 {e(str(user["email"]))}</h2>'
        f'<div class="stats">'
        f'<div class="stat"><div>Баланс</div><div class="v">{e(str(user["balance"]))} ₽</div></div>'
        f'<div class="stat"><div>Реф. код</div><div class="v" style="font-size:16px"><code>{e(str(user["ref_code"]))}</code></div></div>'
        f'<div class="stat"><div>Реферер</div><div class="v" style="font-size:14px">{e(referrer_email)}</div></div>'
        f'<div class="stat"><div>Регистрация</div><div class="v" style="font-size:14px">{e(str(user["created_at"] or ""))}</div></div>'
        f'</div>'
        f'<div class="controls" style="margin-top:12px">'
        f'<a href="{url_for("admin.activate_cabinet_users", bot=bot_id)}"><button type="button" class="gray">← Назад к списку</button></a>'
        f'</div></div>'
    )

    orders_rows = ""
    if orders:
        for o in orders:
            pname = products_dict.get(o.get("product"), o.get("product"))
            orders_rows += f"<tr><td><code>{e(str(o['order_id'] or '')[:16])}...</code></td><td>{e(str(pname))}</td><td>{e(str(o['amount']))} ₽</td><td>{badge_html(str(o['status']))}</td><td class=\"small\">{e(str(o['created_at'] or ''))}</td></tr>"
    else:
        orders_rows = '<tr><td colspan="5">Нет заказов</td></tr>'
    orders_html = f'<div class="box"><h3>🛒 Заказы</h3><table><thead><tr><th>ID</th><th>Продукт</th><th>Сумма</th><th>Статус</th><th>Дата</th></tr></thead><tbody>{orders_rows}</tbody></table></div>'

    ref_rows = ""
    if referrals:
        for r in referrals:
            ref_rows += f"<tr><td>{e(str(r['email']))}</td><td>{badge_html(str(r['status']))}</td><td><strong>+{e(str(r['total_earned']))} ₽</strong></td><td class=\"small\">{e(str(r['created_at'] or ''))}</td></tr>"
    else:
        ref_rows = '<tr><td colspan="4">Нет рефералов</td></tr>'
    referrals_html = f'<div class="box"><h3>👥 Рефералы</h3><table><thead><tr><th>Email</th><th>Статус</th><th>Заработано</th><th>Дата</th></tr></thead><tbody>{ref_rows}</tbody></table></div>'

    tx_rows = ""
    tx_labels = {"referral_bonus": "🎁 Реф. бонус", "withdraw_pending": "📤 Холд", "withdraw_done": "✅ Выплата", "withdraw_rejected": "↩️ Возврат", "bonus": "🎉 Бонус", "correction": "✏️ Корректировка"}
    if transactions:
        for t in transactions:
            label = tx_labels.get(t["tx_type"], t["tx_type"])
            amt = t["amount"]
            amt_class = "free" if amt >= 0 else "used"
            sign = "+" if amt >= 0 else ""
            meta_short = (t.get("meta_json") or "—")[:60]
            tx_rows += f'<tr><td>{e(label)}</td><td><span class="badge {amt_class}">{sign}{e(str(amt))} ₽</span></td><td class="small"><code>{e(meta_short)}</code></td><td class="small">{e(str(t["created_at"] or ""))}</td></tr>'
    else:
        tx_rows = '<tr><td colspan="4">Нет транзакций</td></tr>'
    tx_html = f'<div class="box"><h3>📊 История баланса</h3><table><thead><tr><th>Тип</th><th>Сумма</th><th>Детали</th><th>Дата</th></tr></thead><tbody>{tx_rows}</tbody></table></div>'

    wd_rows = ""
    if withdrawals:
        for w in withdrawals:
            note_short = (w.get("note") or "—")[:50]
            wd_rows += f'<tr><td>#{e(str(w["id"]))}</td><td>{e(str(w["amount"]))} ₽</td><td>{badge_html(str(w["status"]))}</td><td class="small"><code>{e(note_short)}</code></td><td class="small">{e(str(w["created_at"] or ""))}</td><td class="small">{e(str(w["processed_at"] or "—"))}</td></tr>'
    else:
        wd_rows = '<tr><td colspan="6">Нет заявок</td></tr>'
    wd_html = f'<div class="box"><h3>💰 Заявки на вывод</h3><table><thead><tr><th>ID</th><th>Сумма</th><th>Статус</th><th>Реквизиты</th><th>Создана</th><th>Обработана</th></tr></thead><tbody>{wd_rows}</tbody></table></div>'

    actions_html = f"""
    <div class="box">
      <h3>⚙️ Действия</h3>
      <div class="grid2">
        <div>
          <h4 style="margin-bottom:8px">Корректировка баланса</h4>
          <form method="post" class="controls">
            <input type="hidden" name="bot" value="{e(bot_id)}">
            <input type="hidden" name="action" value="adjust_balance">
            <input type="number" name="amount" required placeholder="Сумма (±)" style="width:120px">
            <input type="text" name="note" placeholder="Комментарий">
            <button type="submit" class="green">Применить</button>
          </form>
        </div>
        <div>
          <h4 style="margin-bottom:8px">Управление</h4>
          <div class="controls">
            <form method="post" onsubmit="return confirm('Завершить все сессии?')">
              <input type="hidden" name="bot" value="{e(bot_id)}">
              <input type="hidden" name="action" value="kill_sessions">
              <button type="submit" class="orange">🔑 Сбросить сессии</button>
            </form>
            <form method="post" onsubmit="return confirm('Сбросить пароль? Новый пароль будет показан один раз!')">
              <input type="hidden" name="bot" value="{e(bot_id)}">
              <input type="hidden" name="action" value="reset_password">
              <button type="submit" class="red">🔒 Сбросить пароль</button>
            </form>
          </div>
        </div>
      </div>
    </div>
    """

    html = profile_html + actions_html + orders_html + referrals_html + tx_html + wd_html
    return render_page(html, current_bot=bot_id)


@admin.route("/activate/referral-analytics")
@login_required
def activate_referral_analytics():
    bot_id = request.args.get("bot", "activate")
    if bot_id not in BOTS:
        bot_id = "activate"
    ph = get_placeholder(bot_id)
    page = max(1, safe_int(request.args.get("page", 1)))

    total = db_fetch_val(bot_id, "SELECT COUNT(DISTINCT r.referrer_id) FROM referrals r")
    total_pages = max(1, math.ceil(total / PAGE_SIZE)) if total else 1
    page = max(1, min(page, total_pages))
    offset = (page - 1) * PAGE_SIZE

    rows = db_fetch_all(
        bot_id,
        f"SELECT u.id, u.email, u.balance, COUNT(r.id) AS ref_count, COALESCE(SUM(r.total_earned), 0) AS total_earned "
        f"FROM users u JOIN referrals r ON r.referrer_id = u.id "
        f"GROUP BY u.id, u.email, u.balance ORDER BY total_earned DESC LIMIT {ph} OFFSET {ph}",
        (PAGE_SIZE, offset),
    )

    if rows:
        body_rows = []
        for r in rows:
            body_rows.append(
                f"<tr>"
                f'<td><a href="{url_for("admin.activate_cabinet_user", user_id=r["id"], bot=bot_id)}">{e(str(r["email"]))}</a></td>'
                f"<td><strong>{e(str(r['ref_count']))}</strong></td>"
                f"<td><strong>+{e(str(r['total_earned']))} ₽</strong></td>"
                f"<td>{e(str(r['balance']))} ₽</td>"
                f'<td><a href="{url_for("admin.activate_cabinet_user", user_id=r["id"], bot=bot_id)}"><button type="button" class="gray">Подробнее</button></a></td>'
                f"</tr>"
            )

        pager = '<div class="controls" style="margin-top:12px">'
        if page > 1:
            pager += f'<a href="{url_for("admin.activate_referral_analytics", bot=bot_id, page=page - 1)}"><button type="button" class="gray">← Назад</button></a>'
        pager += f"<span>Страница {page} из {total_pages} | Всего: {total}</span>"
        if page < total_pages:
            pager += f'<a href="{url_for("admin.activate_referral_analytics", bot=bot_id, page=page + 1)}"><button type="button" class="gray">Вперёд →</button></a>'
        pager += "</div>"

        table_html = (
            f'<table><thead><tr><th>Реферер</th><th>Рефералов</th><th>Заработано</th><th>Баланс</th><th></th></tr></thead>'
            f'<tbody>{"".join(body_rows)}</tbody></table>{pager}'
        )
    else:
        table_html = "<p>Нет данных о рефералах</p>"

    html = (
        f'<div class="box"><h2>📊 Реферальная аналитика</h2>'
        f'<div class="controls"><a href="{url_for("admin.index", bot=bot_id)}"><button type="button" class="gray">← Назад</button></a></div></div>'
        f'<div class="box">{table_html}</div>'
    )
    return render_page(html, current_bot=bot_id)

# =========================
# NOTIFY BOT ROUTES
# =========================

@admin.route("/notify/dashboard")
@login_required
def notify_dashboard():
    bot_id = "notify"
    
    # Статистика тихих SKU
    quiet_count = db_fetch_val(bot_id, "SELECT COUNT(*) FROM quiet_skus")
    
    # Статистика followup
    with get_followup_db(bot_id) as conn:
        cur = conn.cursor()
        cur.execute("SELECT COUNT(*) FROM followup_chats WHERE active = 1")
        active_followup = cur.fetchone()[0]
        cur.execute("SELECT COUNT(*) FROM followup_chats WHERE active = 0")
        inactive_followup = cur.fetchone()[0]
    
    # Магазины
    shops_html = "".join([
        f'<div class="stat"><div>{e(s["name"])}</div>'
        f'<div class="small">business_id: {s["business_id"]}</div></div>'
        for s in BOTS[bot_id]["shops"]
    ])
    
    content = f"""
    <div class="box">
      <h2>{e(BOTS[bot_id]["name"])}</h2>
      <p>Управление тихими SKU и follow-up рассылкой</p>
      <div class="controls" style="margin-top:12px">
        <a href="{url_for('admin.notify_quiet_skus')}"><button type="button">📃 Тихие SKU</button></a>
        <a href="{url_for('admin.notify_followup_list')}"><button type="button" class="orange">📃 Заказы в рассылке</button></a>
        <a href="{url_for('admin.notify_chat_management')}"><button type="button" class="blue">💬 Открыть чат / Написать</button></a>
      </div>
      <div class="controls" style="margin-top:8px">
        <a href="{url_for('admin.notify_add_quiet_form')}"><button type="button" class="green">➕ Добавить тихий SKU</button></a>
        <a href="{url_for('admin.notify_followup_add')}"><button type="button" class="green">📨 Добавить в рассылку</button></a>
      </div>
    </div>
    <div class="box">
      <h3>Статистика</h3>
      <div class="stats">
        <div class="stat"><div>Тихих SKU</div><div class="v">{quiet_count}</div></div>
        <div class="stat"><div>В рассылке</div><div class="v">{active_followup}</div><div class="small">активных</div></div>
        <div class="stat"><div>Завершено</div><div class="v">{inactive_followup}</div><div class="small">неактивных</div></div>
      </div>
    </div>
    <div class="box">
      <h3>Магазины</h3>
      <div class="stats">{shops_html}</div>
    </div>
    """
    return render_page(content, current_bot=bot_id)


@admin.route("/notify/quiet-skus")
@login_required
def notify_quiet_skus():
    bot_id = "notify"
    page = max(1, safe_int(request.args.get("page", 1)))
    q = request.args.get("q", "").strip()
    
    # Поиск
    where_clause = ""
    params = []
    if q:
        where_clause = "WHERE sku LIKE ?"
        params.append(f"%{q}%")
    
    total = db_fetch_val(bot_id, f"SELECT COUNT(*) FROM quiet_skus {where_clause}", tuple(params))
    total_pages = max(1, math.ceil(total / PAGE_SIZE)) if total else 1
    page = max(1, min(page, total_pages))
    offset = (page - 1) * PAGE_SIZE
    
    rows = db_fetch_all(
        bot_id,
        f"SELECT sku FROM quiet_skus {where_clause} ORDER BY sku LIMIT {PAGE_SIZE} OFFSET {offset}",
        tuple(params),
    )
    
    # Форма поиска
    search_form = f"""
    <form method="get" class="controls">
      <input type="text" name="q" value="{e(q)}" placeholder="Поиск по SKU">
      <button type="submit">Найти</button>
      <a href="{url_for('admin.notify_quiet_skus')}"><button type="button" class="gray">Сбросить</button></a>
      <a href="{url_for('admin.notify_add_quiet_form')}"><button type="button" class="green">➕ Добавить</button></a>
      <a href="{url_for('admin.notify_dashboard')}"><button type="button" class="gray">← Назад</button></a>
    </form>
    """
    
    if rows:
        body_rows = []
        for r in rows:
            sku = r["sku"]
            del_form = (
                f'<form method="post" action="{url_for("admin.notify_delete_quiet")}" '
                f'onsubmit="return confirm(\'Удалить SKU {e(sku)}?\')" style="display:inline">'
                f'<input type="hidden" name="sku" value="{e(sku)}">'
                f'<button type="submit" class="red">Удалить</button></form>'
            )
            body_rows.append(f"<tr><td><code>{e(sku)}</code></td><td>{del_form}</td></tr>")
        
        # Пагинация
        pager = '<div class="controls" style="margin-top:12px">'
        if page > 1:
            pager += f'<a href="{url_for("admin.notify_quiet_skus", q=q, page=page-1)}"><button type="button" class="gray">← Назад</button></a>'
        pager += f"<span>Страница {page} из {total_pages} | Всего: {total}</span>"
        if page < total_pages:
            pager += f'<a href="{url_for("admin.notify_quiet_skus", q=q, page=page+1)}"><button type="button" class="gray">Вперёд →</button></a>'
        pager += "</div>"
        
        table_html = f"""
        <table>
          <thead><tr><th>SKU</th><th>Действия</th></tr></thead>
          <tbody>{''.join(body_rows)}</tbody>
        </table>
        {pager}
        """
    else:
        table_html = "<p>Нет тихих SKU</p>"
    
    content = f"""
    <div class="box"><h2>📃 Тихие SKU</h2>{search_form}</div>
    <div class="box">{table_html}</div>
    """
    return render_page(content, current_bot=bot_id)


@admin.route("/notify/quiet-skus/add", methods=["GET", "POST"])
@login_required
def notify_add_quiet_form():
    bot_id = "notify"
    
    if request.method == "POST":
        skus_raw = request.form.get("skus", "").strip()
        skus = [s.strip() for s in skus_raw.splitlines() if s.strip()]
        
        added = 0
        for sku in skus:
            try:
                add_quiet_sku_admin(bot_id, sku)
                added += 1
            except Exception:
                pass
        
        flash(f"Добавлено SKU: {added}", "success")
        return redirect(url_for("admin.notify_quiet_skus"))
    
    content = f"""
    <div class="box">
      <h2>➕ Добавить тихие SKU</h2>
      <form method="post">
        <p>Введите SKU (каждый с новой строки):</p>
        <textarea name="skus" required placeholder="MRKT-XXXXX&#10;MRKT-YYYYY" style="height:200px"></textarea>
        <div style="margin-top:12px">
          <button type="submit" class="green">Добавить</button>
          <a href="{url_for('admin.notify_quiet_skus')}"><button type="button" class="gray">Отмена</button></a>
        </div>
      </form>
    </div>
    """
    return render_page(content, current_bot=bot_id)


@admin.route("/notify/quiet-skus/delete", methods=["POST"])
@login_required
def notify_delete_quiet():
    bot_id = "notify"
    sku = request.form.get("sku", "").strip()
    
    if sku:
        if remove_quiet_sku_admin(bot_id, sku):
            flash(f"SKU удалён: {sku}", "success")
        else:
            flash(f"SKU не найден: {sku}", "error")
    
    return redirect(url_for("admin.notify_quiet_skus"))


@admin.route("/notify/followup")
@login_required
def notify_followup_list():
    bot_id = "notify"
    show_all = request.args.get("all", "").strip() == "1"
    
    rows = list_followup_orders_admin(bot_id, active_only=not show_all)
    shops = {s["business_id"]: s["name"] for s in BOTS[bot_id]["shops"]}
    
    # Фильтр
    filter_btns = f"""
    <div class="filter-bar">
      <span class="label">Показать:</span>
      <a class="{'active' if not show_all else ''}" href="{url_for('admin.notify_followup_list')}">Только активные</a>
      <a class="{'active' if show_all else ''}" href="{url_for('admin.notify_followup_list', all='1')}">Все</a>
      <a href="{url_for('admin.notify_dashboard')}"><button type="button" class="gray">← Назад</button></a>
    </div>
    """
    
    if rows:
        body_rows = []
        for r in rows:
            shop_name = shops.get(r["business_id"], f"ID:{r['business_id']}")
            status_badge = badge_html("active") if r["active"] == 1 else badge_html("blocked")
            
            # Кнопка "Убрать из рассылки" только для активных
            actions = ""
            if r["active"] == 1:
                actions = (
                    f'<form method="post" action="{url_for("admin.notify_followup_remove")}" '
                    f'onsubmit="return confirm(\'Убрать заказ {r["order_id"]} из рассылки?\')" style="display:inline">'
                    f'<input type="hidden" name="order_id" value="{r["order_id"]}">'
                    f'<button type="submit" class="red">🛑 Убрать</button></form>'
                )
            
            body_rows.append(
                f"<tr>"
                f"<td><strong>{r['order_id']}</strong></td>"
                f"<td>{e(shop_name)}</td>"
                f"<td>{r['chat_id']}</td>"
                f"<td>Вариант {r['variant']}</td>"
                f"<td>{status_badge}</td>"
                f"<td class='small'>{e(str(r.get('last_checked_at') or '—'))}</td>"
                f"<td class='small'>{e(str(r.get('last_sent_date') or '—'))}</td>"
                f"<td>{actions}</td>"
                f"</tr>"
            )
        
        table_html = f"""
        <table>
          <thead>
            <tr>
              <th>Заказ</th><th>Магазин</th><th>Chat ID</th><th>Вариант</th>
              <th>Статус</th><th>Проверен</th><th>Отправлено</th><th>Действия</th>
            </tr>
          </thead>
          <tbody>{''.join(body_rows)}</tbody>
        </table>
        """
    else:
        table_html = "<p>Нет заказов в рассылке</p>"
    
    content = f"""
    <div class="box">
      <h2>📨 Заказы в рассылке (Follow-up)</h2>
      {filter_btns}
    </div>
    <div class="box">{table_html}</div>
    """
    return render_page(content, current_bot=bot_id)


@admin.route("/notify/followup/remove", methods=["POST"])
@login_required
def notify_followup_remove():
    bot_id = "notify"
    order_id = safe_int(request.form.get("order_id", 0), 0)
    
    if order_id:
        if deactivate_followup_order_admin(bot_id, order_id):
            flash(f"Заказ {order_id} убран из рассылки", "success")
        else:
            flash(f"Заказ {order_id} не найден в рассылке", "error")
    
    return redirect(url_for("admin.notify_followup_list"))

@admin.route("/notify/followup/add", methods=["GET", "POST"])
@login_required
def notify_followup_add():
    bot_id = "notify"
    if request.method == "POST":
        order_id = safe_int(request.form.get("order_id", 0), 0)
        variant = safe_int(request.form.get("variant", 1), 1)
        
        if not order_id:
            flash("Введите корректный номер заказа", "error")
            return redirect(url_for("admin.notify_followup_add"))
            
        shop, _ = notify_find_shop_for_order(order_id)
        if not shop:
            flash(f"Заказ {order_id} не найден ни в одном магазине", "error")
            return redirect(url_for("admin.notify_followup_add"))
            
        chat_id, last_message_id = notify_create_chat_and_get_last_msg(shop, order_id)
        if not chat_id:
            flash("Не удалось создать/найти чат для этого заказа", "error")
            return redirect(url_for("admin.notify_followup_add"))
            
        with get_followup_db(bot_id) as conn:
            cur = conn.cursor()
            cur.execute(
                """
                INSERT OR REPLACE INTO followup_chats
                (order_id, business_id, campaign_id, chat_id, variant, active, last_message_id)
                VALUES (?, ?, ?, ?, ?, 1, ?)
                """,
                (order_id, shop["business_id"], shop["campaign_id"], chat_id, variant, last_message_id)
            )
            conn.commit()
            
        flash(f"Заказ {order_id} добавлен в рассылку (Вариант {variant})", "success")
        return redirect(url_for("admin.notify_followup_list"))

    content = f"""
    <div class="box">
      <h2>📨 Добавить заказ в рассылку</h2>
      <form method="post">
        <div class="controls">
          <input type="number" name="order_id" required placeholder="Номер заказа (ID)">
          <select name="variant">
            <option value="1">Вариант текста 1</option>
            <option value="2">Вариант текста 2</option>
          </select>
          <button type="submit" class="green">Добавить</button>
        </div>
      </form>
      <div style="margin-top:12px">
        <a href="{url_for('admin.notify_dashboard')}"><button type="button" class="gray">← Назад</button></a>
      </div>
    </div>
    """
    return render_page(content, current_bot=bot_id)


@admin.route("/notify/chat", methods=["GET", "POST"])
@login_required
def notify_chat_management():
    bot_id = "notify"
    result_html = ""
    
    if request.method == "POST":
        order_id = safe_int(request.form.get("order_id", 0), 0)
        message_text = request.form.get("message_text", "").strip()
        
        if not order_id:
            flash("Введите корректный номер заказа", "error")
        else:
            shop, _ = notify_find_shop_for_order(order_id)
            if not shop:
                flash(f"Заказ {order_id} не найден", "error")
            else:
                chat_id, _ = notify_create_chat_and_get_last_msg(shop, order_id)
                if not chat_id:
                    flash("Не удалось открыть чат", "error")
                else:
                    # Чат найден
                    if message_text:
                        # Отправка сообщения
                        success, resp = notify_send_message_to_market(shop, chat_id, message_text)
                        if success:
                            flash(f"Сообщение успешно отправлено в чат заказа {order_id}!", "success")
                        else:
                            flash(f"Ошибка отправки: {resp}", "error")
                    
                    # Показываем инфу о чате
                    result_html = f"""
                    <div class="box">
                      <h3>Информация о чате</h3>
                      <p><strong>Магазин:</strong> {e(shop['name'])}</p>
                      <p><strong>Заказ ID:</strong> {order_id}</p>
                      <p><strong>Чат ID:</strong> <code>{chat_id}</code></p>
                    </div>
                    """

    content = f"""
    <div class="box">
      <h2>💬 Управление чатами Маркета</h2>
      <p>Здесь можно открыть чат (узнать его ID) или сразу отправить сообщение клиенту.</p>
      
      <form method="post" style="margin-top:16px;">
        <div class="controls">
          <input type="number" name="order_id" required placeholder="Номер заказа (ID)">
        </div>
        <div style="margin-top:12px;">
          <textarea name="message_text" placeholder="Текст сообщения (оставь пустым, если нужно только открыть чат)" style="height:100px; width:100%; max-width:600px;"></textarea>
        </div>
        <div class="controls" style="margin-top:12px;">
          <button type="submit" class="blue">🚀 Выполнить</button>
          <a href="{url_for('admin.notify_dashboard')}"><button type="button" class="gray">Отмена</button></a>
        </div>
      </form>
    </div>
    {result_html}
    """
    return render_page(content, current_bot=bot_id)


# =========================
# REGISTER BLUEPRINT & ROOT
# =========================
app.register_blueprint(admin)


@app.route("/")
def root():
    return redirect(url_for("admin.index"))


if __name__ == "__main__":
    app.run(host="0.0.0.0", port=9000, debug=False)